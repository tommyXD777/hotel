from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash, send_file
from datetime import datetime, timedelta, date, timezone, time
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from werkzeug.security import check_password_hash
import pymysql
import threading
import time
import pytz
from zoneinfo import ZoneInfo
import _thread

bogota = pytz.timezone("America/Bogota")
now = datetime.now(bogota)

print("Bogotá:", now)

def require_user_session_json():
    """Versión para rutas que devuelven JSON."""
    user_id = session.get('usuario_id')
    if not user_id:
        return None
    return user_id


app = Flask(__name__)
app.secret_key = "una_clave_muy_secreta_y_larga"  # 🔑 obligatorio para sesión y flash

# ----------------- CONEXIÓN DB -----------------
def get_db_connection():
    """Obtiene una conexión a la base de datos con manejo de errores mejorado"""
    try:
        print("Intentando conectar a MySQL...")
        print(f"   Host: localhost")
        print(f"   Usuario: nelson")
        print(f"   Puerto: 3311")  # actualizado el puerto mostrado
        print(f"   Base de datos: bd_hostal")

        conn = pymysql.connect(
            host='mysql',  # Asegurándonos que sea localhost
            user='nelson',
            password='3011551141.Arias',
            database='bd_hostal',
            charset='utf8mb4',
            autocommit=False,
            port=3306  # cambiado de 3306 a 3311
        )
        print("Conexion exitosa a MySQL", flush=True)
        return conn
    except pymysql.err.OperationalError as e:
        error_code = e.args[0]
        if error_code == 2003:
            print("Error 2003: No se puede conectar al servidor MySQL", flush=True)
            print("Soluciones posibles:")
            print("   1. Verifica que MySQL este ejecutandose: 'net start mysql' (Windows)")
            print("   2. Verifica que MySQL este en el puerto 3311")  # actualizado el puerto en el mensaje
            print("   3. Intenta conectarte manualmente: mysql -u nelson -p -P 3311")  # agregado -P 3311
        elif error_code == 1045:
            print("Error 1045: Acceso denegado - credenciales incorrectas")
            print("Verifica usuario y contrasena en MySQL")
        elif error_code == 1049:
            print("Error 1049: Base de datos 'bd_hostal' no existe")
            print("Crea la base de datos: CREATE DATABASE bd_hostal;")
        else:
            print(f"Error MySQL {error_code}: {e}")
        return None
    except Exception as e:
        print(f"Error inesperado de conexion: {e}")
        return None

def get_current_user_id():
    """Obtiene el ID del usuario actual desde la sesión"""
    return session.get('usuario_id')

def require_user_session():
    """Verifica que el usuario esté logueado y retorna su ID"""
    user_id = get_current_user_id()
    if not user_id:
        flash("Debes iniciar sesión para acceder a esta función", "error")
        return None
    return user_id

# ----------------- LOGIN -----------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('login.html')

    if request.is_json:
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '')
    else:
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

    # Validación de campos vacíos
    if not username or not password:
        error_msg = "Usuario y contraseña son requeridos"
        if request.is_json:
            return jsonify({"success": False, "error": error_msg})
        else:
            flash(error_msg, "error")
            return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        error_msg = "Error de conexión a la base de datos. Verifica que MySQL esté ejecutándose."
        if request.is_json:
            return jsonify({"success": False, "error": error_msg})
        else:
            flash(error_msg, "error")
            return redirect(url_for('login'))

    try:
        cur = conn.cursor()
        cur.execute("SELECT id, username, password FROM usuarios WHERE username = %s", (username,))
        user = cur.fetchone()
        
        if user and check_password_hash(user[2], password):
            session['usuario_id'] = user[0]
            session['usuario'] = user[1]
            if request.is_json:
                return jsonify({"success": True, "redirect": url_for('index')})
            else:
                flash(f"Bienvenido, {username}!", "success")
                return redirect(url_for('index'))
        else:
            error_msg = "Usuario o contraseña incorrectos"
            if request.is_json:
                return jsonify({"success": False, "error": error_msg})
            else:
                flash(error_msg, "error")
                return redirect(url_for('login'))
            
    except pymysql.MySQLError as e:
        print(f"Error MySQL en login: {e}")
        error_msg = "Error en el sistema de autenticación"
        if request.is_json:
            return jsonify({"success": False, "error": error_msg})
        else:
            flash(error_msg, "error")
            return redirect(url_for('login'))
    except Exception as e:
        print(f"Error inesperado en login: {e}")
        error_msg = "Error interno del servidor"
        if request.is_json:
            return jsonify({"success": False, "error": error_msg})
        else:
            flash(error_msg, "error")
            return redirect(url_for('login'))
    finally:
        if 'cur' in locals():
            cur.close()
        if conn:
            conn.close()

# ----------------- LOGOUT -----------------
@app.route('/logout', methods=['POST'])
def logout():
    session.clear()  # limpia la sesión
    flash("Sesión cerrada con éxito", "success")
    return redirect(url_for('login'))

# ----------------- PROTECCIÓN DE RUTAS -----------------
@app.before_request
def require_login():
    rutas_libres = {'login', 'static', 'test_db'} # Added test_db to free routes
    if request.endpoint not in rutas_libres and 'usuario_id' not in session:
        # Prevent redirect loops for /login itself
        if request.path != url_for('login'):
            return redirect(url_for('login'))

@app.route('/test-db')
def test_db():
    """Ruta para probar la conexión a la base de datos con diagnóstico detallado"""
    print("🧪 Ejecutando prueba de conexión...")
    
    try:
        import pymysql
        print("✅ Módulo pymysql importado correctamente")
    except ImportError:
        return jsonify({"success": False, "error": "pymysql no está instalado. Ejecuta: pip install pymysql"})
    
    conn = get_db_connection()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("SELECT VERSION()")
            version = cur.fetchone()
            cur.execute("SELECT DATABASE()")
            database = cur.fetchone()
            cur.close()
            conn.close()
            
            return jsonify({
                "success": True, 
                "message": "✅ Conexión a base de datos exitosa",
                "mysql_version": version[0] if version else "Desconocida",
                "database": database[0] if database else "Desconocida"
            })
        except Exception as e:
            return jsonify({"success": False, "error": f"Error en query: {str(e)}"})
    else:
        return jsonify({
            "success": False, 
            "error": "❌ No se pudo conectar a la base de datos. Revisa la consola para más detalles."
        })

# ----------------- INDEX -----------------
@app.route('/')
def index():
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))

    print(f"[DEBUG] index - Inicio. User ID: {user_id}")

    conn = get_db_connection()
    if not conn:
        print("[DEBUG] index - Error de conexión a la base de datos")
        flash("Error de conexión a la base de datos")
        return render_template('index.html', habitaciones=[], rooms=[])

    try:
        cur = conn.cursor()

        cur.execute("SELECT id, numero, descripcion, estado FROM habitaciones WHERE usuario_id = %s ORDER BY numero", (user_id,))
        habitaciones_db = cur.fetchall()
        print(f"[DEBUG] index - Habitaciones encontradas: {len(habitaciones_db)}")

        rooms = []
        fecha_actual = datetime.now()

        for h in habitaciones_db:
            room_id, numero, descripcion, estado = h
            print(f"[DEBUG] index - Procesando habitación {numero} (ID: {room_id}), estado: {estado}")

            # Ordena: primero el cliente principal (valor > 0), luego acompañantes
            cur.execute("""SELECT nombre, telefono, observacion, check_out, id, check_in, valor, tipo_doc, numero_doc, procedencia FROM clientes WHERE habitacion_id = %s AND (check_out IS NULL OR check_out > NOW()) ORDER BY (valor > 0) DESC, check_in DESC""", (room_id,))
            clientes = cur.fetchall()
            print(f"[DEBUG] index - Clientes encontrados para habitación {numero}: {len(clientes)}")

            inquilino_principal = clientes[0][0] if clientes else None
            telefono = clientes[0][1] if clientes else None
            observacion = clientes[0][2] if clientes else None
            fecha_salida = clientes[0][3] if clientes and clientes[0][3] else None
            cliente_id = clientes[0][4] if clientes else None
            fecha_ingreso = clientes[0][5] if clientes and clientes[0][5] else None
            valor = clientes[0][6] if clientes and clientes[0][6] else None
            tipo_doc = clientes[0][7] if clientes else None
            numero_doc = clientes[0][8] if clientes else None
            procedencia = clientes[0][9] if clientes and clientes[0][9] else None

            dias_ocupada = 0
            if fecha_ingreso and fecha_salida:
                # FIX: Calculate days correctly - check_out is at 1 PM, so we need to consider the full days
                # If check-in is 2025-10-23 14:00 and check-out is 2025-10-25 13:00, that's 2 full days
                dias_ocupada = (fecha_salida.date() - fecha_ingreso.date()).days
                if dias_ocupada <= 0:
                    dias_ocupada = 1
            elif fecha_ingreso:
                dias_ocupada = (datetime.now().date() - fecha_ingreso.date()).days + 1

            print(f"[DEBUG] index - Habitación {numero}: inquilino_principal={inquilino_principal}, fecha_ingreso={fecha_ingreso}, fecha_salida={fecha_salida}, valor={valor}, dias_ocupada={dias_ocupada}")
            print(f"[DEBUG] index - Estado final de habitación {numero}: {estado}")

            current_time = datetime.now()

            if fecha_salida and fecha_salida <= current_time:
                if estado in ['ocupada', 'reservado']:
                    print(f"[DEBUG] index - Habitación {numero} expirada, cambiando a 'libre'")
                    # Update room to libre (green) when checkout time has passed
                    cur.execute("UPDATE habitaciones SET estado = 'libre' WHERE id = %s AND usuario_id = %s", (room_id, user_id))
                    estado = 'libre'
                    # Clear expired client data
                    cur.execute("UPDATE clientes SET check_out = NOW() WHERE habitacion_id = %s AND check_out > NOW()", (room_id,))
                    conn.commit()
                    # Reset client data for display
                    inquilino_principal = None
                    telefono = None
                    observacion = None
                    fecha_salida = None
                    cliente_id = None
                    fecha_ingreso = None
                    valor = None
                    tipo_doc = None
                    numero_doc = None
                    procedencia = None
                    dias_ocupada = 0

            elif estado == 'libre' and clientes:
                print(f"[DEBUG] index - Habitación {numero} está libre pero tiene clientes, verificando reservas")
                # Si hay clientes pero la habitación está libre, verificar si es una reserva que debe activarse
                for cliente in clientes:
                    fecha_ingreso_cliente = cliente[5]
                    if fecha_ingreso_cliente and fecha_ingreso_cliente.date() <= fecha_actual.date():
                        print(f"[DEBUG] index - Activando reserva para habitación {numero}")
                        # La reserva debe activarse
                        cur.execute("UPDATE habitaciones SET estado = 'ocupada' WHERE id = %s AND usuario_id = %s", (room_id, user_id))
                        estado = 'ocupada'
                        conn.commit()
                        break

            if estado == 'ocupada' and dias_ocupada >= 30:
                print(f"[DEBUG] index - Habitación {numero} cambiando a mensualidad (dias_ocupada={dias_ocupada})")
                cur.execute("UPDATE habitaciones SET estado = 'mensualidad' WHERE id = %s AND usuario_id = %s", (room_id, user_id))
                estado = 'mensualidad'
                conn.commit()

            rooms.append({
                "id": room_id,
                "numero": numero,
                "descripcion": descripcion,
                "estado": estado,
                "inquilino_principal": inquilino_principal,
                "telefono": telefono,
                "observacion": observacion,
                "fecha_salida": fecha_salida,
                "fecha_ingreso": fecha_ingreso,
                "valor": valor,
                "dias_ocupada": dias_ocupada,
                "num_personas_ocupadas": len(clientes) if clientes else 0,
                "personas_list": [
                    {
                        "nombre": c[0],
                        "telefono": c[1],
                        "id": c[4],
                        "tipo_doc": c[7],
                        "numero_doc": c[8],
                        "procedencia": c[9]
                    } for c in clientes
                ] if clientes else [],
                "cliente_id": cliente_id,
                "tipo_doc": tipo_doc,
                "numero_doc": numero_doc,
                "procedencia": procedencia
            })

        print(f"[DEBUG] index - Total habitaciones procesadas: {len(rooms)}")
        return render_template('index.html', habitaciones=habitaciones_db, rooms=rooms)

    except pymysql.MySQLError as e:
        print(f"[DEBUG] index - Error MySQL: {str(e)}")
        flash(f"Error en la base de datos: {str(e)}")
        return render_template('index.html', habitaciones=[], rooms=[])
    finally:
        if 'cur' in locals():
            cur.close()
        if conn:
            conn.close()

# ----------------- REGISTRAR -----------------
@app.route('/registrar', methods=['POST'])
def registrar():
    flash("Reserva registrada con éxito")
    return redirect(url_for('index'))

# ----------------- GUARDAR CLIENTE -----------------
@app.route('/guardar_cliente', methods=['POST'])
def guardar_cliente():
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))
    
    habitacion_id = request.form['habitacion_id']
    nombre = request.form['nombre']
    telefono = request.form.get('telefono')
    observacion = request.form.get('observacion')
    check_in = request.form['check_in']
    check_out = request.form['check_out']
    valor = request.form['valor']
    tipo_doc = request.form.get('tipo_doc', 'C.c')
    numero_doc = request.form.get('numero_doc', '')
    procedencia = request.form.get('procedencia', '')

    conn = get_db_connection()
    if not conn:
        flash('Error de conexión a la base de datos.', 'error')
        return redirect(url_for('index'))

    try:
        cur = conn.cursor()
        
        cur.execute("SELECT id FROM habitaciones WHERE id = %s AND usuario_id = %s", (habitacion_id, user_id))
        if not cur.fetchone():
            flash('No tienes permisos para modificar esta habitación.', 'error')
            return redirect(url_for('index'))
        
        cur.execute("""SELECT COUNT(*) FROM clientes WHERE habitacion_id = %s AND (check_out IS NULL OR check_out > NOW())""", (habitacion_id,))
        clientes_actuales = cur.fetchone()[0]

        if clientes_actuales >= 4:
            flash('La habitación ha alcanzado el límite máximo de 4 clientes.', 'error')
            return redirect(url_for('index'))

        cur.execute("""INSERT INTO clientes (habitacion_id, nombre, telefono, observacion, check_in, check_out, valor, tipo_doc, numero_doc, procedencia) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (habitacion_id, nombre, telefono, observacion, check_in, check_out, valor, tipo_doc, numero_doc, procedencia))

        cur.execute("UPDATE habitaciones SET estado = 'ocupada' WHERE id = %s AND usuario_id = %s""", (habitacion_id, user_id))

        conn.commit()
        flash('Cliente registrado y habitación marcada como ocupada.', 'success')

    except pymysql.MySQLError as e:
        conn.rollback()
        flash(f'Error al registrar cliente: {e}', 'danger')
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('index'))

# ----------------- EDITAR CLIENTE -----------------
@app.route('/editar_cliente/<int:cliente_id>')
def editar_cliente(cliente_id):
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión a la base de datos.', 'error')
        return redirect(url_for('index'))

    try:
        cur = conn.cursor()
        cur.execute("""SELECT c.id, c.nombre, c.tipo_doc, c.numero_doc, c.telefono, c.procedencia, c.check_in, c.check_out, c.valor, c.observacion, c.habitacion_id FROM clientes c JOIN habitaciones h ON c.habitacion_id = h.id WHERE c.id = %s AND h.usuario_id = %s""", (cliente_id, user_id))
        cliente = cur.fetchone()

        if not cliente:
            flash('Cliente no encontrado o no tienes permisos para editarlo.', 'error')
            return redirect(url_for('index'))

        cur.execute("""SELECT id, numero, descripcion, estado FROM habitaciones WHERE usuario_id = %s AND (estado = 'libre' OR id = %s) ORDER BY numero""", (user_id, cliente[10]))
        habitaciones_disponibles = cur.fetchall()

        return render_template('editar_cliente.html', cliente=cliente, habitaciones_disponibles=habitaciones_disponibles)

    except pymysql.MySQLError as e:
        flash(f'Error en la base de datos: {str(e)}', 'error')
        return redirect(url_for('index'))
    finally:
        cur.close()
        conn.close()

# ----------------- ACTUALIZAR CLIENTE -----------------
@app.route('/actualizar_cliente', methods=['POST'])
def actualizar_cliente():
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))
    
    try:
        cliente_id_str = request.form.get('cliente_id', '').strip()
        if not cliente_id_str:
            flash('ID de cliente no válido.', 'error')
            return redirect(url_for('index'))
        
        cliente_id = int(cliente_id_str)
        nombre = request.form['nombre']
        tipo_doc = request.form['tipo_doc']
        numero_doc = request.form['numero_doc']
        telefono = request.form['telefono']
        procedencia = request.form['procedencia']
        nueva_habitacion_id = request.form.get('habitacion_id', '').strip()

        conn = get_db_connection()
        if not conn:
            flash("Error de conexión a la base de datos", "error")
            return redirect(url_for('index'))

        cur = conn.cursor()
        
        cur.execute("""
            SELECT c.habitacion_id, c.check_in 
            FROM clientes c 
            JOIN habitaciones h ON c.habitacion_id = h.id 
            WHERE c.id = %s AND h.usuario_id = %s
        """, (cliente_id, user_id))
        cliente_actual = cur.fetchone()
        
        if not cliente_actual:
            flash('No tienes permisos para editar este cliente.', 'error')
            return redirect(url_for('index'))
        
        habitacion_actual_id = cliente_actual[0]
        check_in_cliente = cliente_actual[1]
        
        # Si se está cambiando de habitación
        if nueva_habitacion_id and nueva_habitacion_id != str(habitacion_actual_id):
            nueva_habitacion_id = int(nueva_habitacion_id)
            
            # Verificar que la nueva habitación esté disponible
            cur.execute("SELECT estado FROM habitaciones WHERE id = %s AND usuario_id = %s", (nueva_habitacion_id, user_id))
            nueva_habitacion = cur.fetchone()
            
            if not nueva_habitacion:
                flash('La habitación seleccionada no existe o no tienes permisos.', 'error')
                return redirect(url_for('editar_cliente', cliente_id=cliente_id))
            
            if nueva_habitacion[0] != 'libre':
                flash('La habitación seleccionada no está disponible.', 'error')
                return redirect(url_for('editar_cliente', cliente_id=cliente_id))
            
            cur.execute("""
                SELECT id FROM clientes 
                WHERE habitacion_id = %s 
                AND check_in = %s 
                AND (check_out IS NULL OR check_out > NOW())
            """, (habitacion_actual_id, check_in_cliente))
            huespedes_misma_estadia = cur.fetchall()
            
            for huesped in huespedes_misma_estadia:
                cur.execute("UPDATE clientes SET habitacion_id = %s WHERE id = %s", (nueva_habitacion_id, huesped[0]))
            
            # Actualizar estados de habitaciones
            cur.execute("UPDATE habitaciones SET estado = 'ocupada' WHERE id = %s", (nueva_habitacion_id,))
            
            # Verificar si la habitación anterior debe liberarse
            cur.execute("""
                SELECT COUNT(*) FROM clientes 
                WHERE habitacion_id = %s 
                AND (check_out IS NULL OR check_out > NOW())
            """, (habitacion_actual_id,))
            clientes_restantes = cur.fetchone()[0]
            
            if clientes_restantes == 0:
                cur.execute("UPDATE habitaciones SET estado = 'libre' WHERE id = %s", (habitacion_actual_id,))
            
            flash(f'Se movieron {len(huespedes_misma_estadia)} huésped(es) a la nueva habitación.', 'success')
        
        # Actualizar datos del cliente específico
        cur.execute("""
            UPDATE clientes 
            SET nombre = %s, tipo_doc = %s, numero_doc = %s, telefono = %s, procedencia = %s 
            WHERE id = %s
        """, (nombre, tipo_doc, numero_doc, telefono, procedencia, cliente_id))
        
        conn.commit()
        flash('Cliente actualizado exitosamente.', 'success')
        return redirect(url_for('index'))

    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
        flash(f'Error al actualizar cliente: {str(e)}', 'error')
        return redirect(url_for('index'))
    finally:
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()

# ----------------- AGREGAR CLIENTE HABITACIÓN -----------------
@app.route('/agregar_cliente_habitacion/<int:habitacion_id>')
def agregar_cliente_habitacion(habitacion_id):
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión a la base de datos.', 'error')
        return redirect(url_for('index'))

    try:
        cur = conn.cursor()
        cur.execute("SELECT numero, descripcion FROM habitaciones WHERE id = %s AND usuario_id = %s", (habitacion_id, user_id))
        habitacion = cur.fetchone()

        if not habitacion:
            flash('Habitación no encontrada o no tienes permisos para acceder a ella.', 'error')
            return redirect(url_for('index'))

        cur.execute("""SELECT COUNT(*) FROM clientes WHERE habitacion_id = %s AND (check_out IS NULL OR check_out > NOW())""", (habitacion_id,))
        clientes_actuales = cur.fetchone()[0]

        cur.execute("""SELECT check_in, check_out, valor, observacion FROM clientes WHERE habitacion_id = %s AND (check_out IS NULL OR check_out > NOW()) LIMIT 1""", (habitacion_id,))
        datos_referencia = cur.fetchone()

        if clientes_actuales >= 4:
            flash(f'La habitación {habitacion[1]} ya está en su capacidad máxima (4 personas).', 'error')
            return redirect(url_for('index'))

        return render_template('agregar_cliente.html', habitacion=habitacion, habitacion_id=habitacion_id, datos_referencia=datos_referencia)

    except pymysql.MySQLError as e:
        flash(f'Error en la base de datos: {str(e)}', 'error')
        return redirect(url_for('index'))
    finally:
        cur.close()
        conn.close()

# ----------------- GUARDAR NUEVO CLIENTE -----------------
@app.route('/guardar_nuevo_cliente', methods=['POST'])
def guardar_nuevo_cliente():
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    try:
        print(f"[DEBUG] guardar_nuevo_cliente - Inicio. User ID: {user_id}")

        if request.is_json:
            data = request.get_json()
            habitacion_id_str = str(data.get('habitacion_id', '')).strip()
            if not habitacion_id_str:
                print("[DEBUG] guardar_nuevo_cliente - ID de habitación no válido")
                return jsonify({"success": False, "error": "ID de habitación no válido"})
            habitacion_id = int(habitacion_id_str)

            nombre = data['nombre']
            tipo_doc = data['tipo_doc']
            numero_doc = data['numero_doc']
            telefono = data['telefono']
            procedencia = data['procedencia']
            check_in = data['check_in']
            check_out_fecha = data['check_out_fecha']
            valor = data['valor']
            observacion = data['observacion']
            personas_adicionales = data.get('personas_adicionales', [])
        else:
            habitacion_id_str = request.form.get('habitacion_id', '').strip()
            if not habitacion_id_str:
                print("[DEBUG] guardar_nuevo_cliente - ID de habitación no válido (form)")
                flash('ID de habitación no válido.', 'error')
                return redirect(url_for('index'))
            habitacion_id = int(habitacion_id_str)

            nombre = request.form['nombre']
            tipo_doc = request.form['tipo_doc']
            numero_doc = request.form['numero_doc']
            telefono = request.form['telefono']
            procedencia = request.form['procedencia']
            check_in = request.form['check_in']
            check_out_fecha = request.form['check_out_fecha']
            valor = request.form['valor']
            observacion = request.form['observacion']
            personas_adicionales = []

        print(f"[DEBUG] guardar_nuevo_cliente - Datos recibidos: habitacion_id={habitacion_id}, nombre={nombre}, check_in={check_in}, check_out_fecha={check_out_fecha}, valor={valor}")

        check_in_dt = datetime.strptime(check_in, "%Y-%m-%dT%H:%M")
        check_out_dt = datetime.strptime(check_out_fecha, "%Y-%m-%d")

        # The previous validation prevented creating reservations where checkout was in the past
        # Now we only validate that checkout is after check-in, which is the correct business logic
        
        conn = get_db_connection()
        if not conn:
            print("[DEBUG] guardar_nuevo_cliente - Error de conexión a la base de datos")
            if request.is_json:
                return jsonify({"success": False, "error": "Error de conexión a la base de datos"})
            flash('Error de conexión a la base de datos.', 'error')
            return redirect(url_for('index'))

        cur = conn.cursor()
        
        # Get user's configured checkout time
        cur.execute("SELECT checkout_hora FROM usuarios WHERE id = %s", (user_id,))
        checkout_config = cur.fetchone()
        
        if checkout_config and checkout_config[0]:
            checkout_time = checkout_config[0]
            
            # Handle different types returned by MySQL
            if isinstance(checkout_time, timedelta):
                # MySQL TIME type returns as timedelta
                total_seconds = int(checkout_time.total_seconds())
                checkout_hour = total_seconds // 3600
                checkout_minute = (total_seconds % 3600) // 60
            elif isinstance(checkout_time, str):
                # String format "HH:MM:SS" or "HH:MM"
                checkout_time_obj = datetime.strptime(checkout_time[:5], "%H:%M").time()
                checkout_hour = checkout_time_obj.hour
                checkout_minute = checkout_time_obj.minute
            elif isinstance(checkout_time, time):
                # Already a time object
                checkout_hour = checkout_time.hour
                checkout_minute = checkout_time.minute
            else:
                # Fallback to default
                checkout_hour = 13
                checkout_minute = 0
        else:
            # Default to 1 PM if not configured
            checkout_hour = 13
            checkout_minute = 0
        
        check_out_dt = datetime(check_out_dt.year, check_out_dt.month, check_out_dt.day, checkout_hour, checkout_minute)
        print(f"[DEBUG] guardar_nuevo_cliente - Using checkout time: {checkout_hour}:{checkout_minute:02d}")

        if check_out_dt <= check_in_dt:
            if request.is_json:
                return jsonify({"success": False, "error": "La fecha de check-out debe ser posterior a la fecha de check-in"})
            flash('La fecha de check-out debe ser posterior a la fecha de check-in.', 'error')
            return redirect(url_for('index'))

        hora_ingreso_time = check_in_dt.time()
        print(f"[DEBUG] guardar_nuevo_cliente - Fechas procesadas: check_in_dt={check_in_dt}, check_out_dt={check_out_dt}, hora_ingreso={hora_ingreso_time}")

        cur.execute("SELECT id FROM habitaciones WHERE id = %s AND usuario_id = %s", (habitacion_id, user_id))
        habitacion_check = cur.fetchone()
        if not habitacion_check:
            print(f"[DEBUG] guardar_nuevo_cliente - No tiene permisos para habitación {habitacion_id}")
            if request.is_json:
                return jsonify({"success": False, "error": "No tienes permisos para modificar esta habitación"})
            flash('No tienes permisos para modificar esta habitación.', 'error')
            return redirect(url_for('index'))

        reserva_existente = verificar_reserva_existente(habitacion_id, user_id)
        print(f"[DEBUG] guardar_nuevo_cliente - Reserva existente: {reserva_existente}")

        if reserva_existente:
            print("[DEBUG] guardar_nuevo_cliente - Actualizando reserva existente")
            datos_cliente = {
                'nombre': nombre,
                'tipo_doc': tipo_doc,
                'numero_doc': numero_doc,
                'telefono': telefono,
                'procedencia': procedencia,
                'check_in': check_in_dt,
                'check_out': check_out_dt,
                'valor': valor,
                'observacion': observacion
            }

            if actualizar_reserva_existente(reserva_existente[0], datos_cliente):
                print("[DEBUG] guardar_nuevo_cliente - Reserva existente actualizada, agregando personas adicionales")
                # Add additional persons if any
                for persona in personas_adicionales:
                    if persona.get('nombre'):
                        cur.execute("""
                            INSERT INTO clientes (habitacion_id, nombre, tipo_doc, numero_doc, telefono, procedencia, check_in, check_out, valor, observacion)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (habitacion_id, persona['nombre'], persona.get('tipo_doc', 'C.c'), persona.get('numero_doc', ''),
                              persona.get('telefono', ''), persona.get('procedencia', ''), check_in_dt, check_out_dt, 0, ''))
                        print(f"[DEBUG] guardar_nuevo_cliente - Persona adicional agregada: {persona['nombre']}")

                cur.execute("""UPDATE habitaciones SET estado = 'ocupada' WHERE id = %s AND usuario_id = %s""", (habitacion_id, user_id))
                conn.commit()
                print(f"[DEBUG] guardar_nuevo_cliente - Habitación {habitacion_id} actualizada a 'ocupada'")

                if request.is_json:
                    return jsonify({"success": True, "message": "Reserva completada exitosamente. La habitación ahora está ocupada."})
                else:
                    flash('Reserva actualizada exitosamente.', 'success')
                    return redirect(url_for('index'))
            else:
                print("[DEBUG] guardar_nuevo_cliente - Error al actualizar reserva existente")
                if request.is_json:
                    return jsonify({"success": False, "error": "Error al actualizar la reserva existente"})
                flash('Error al actualizar la reserva.', 'error')
                return redirect(url_for('index'))
        else:
            print("[DEBUG] guardar_nuevo_cliente - Creando nueva reserva")
            # Create new reservation if none exists
            cur.execute("""SELECT COUNT(*) FROM clientes WHERE habitacion_id = %s AND (check_out IS NULL OR check_out > NOW())""", (habitacion_id,))
            clientes_actuales = cur.fetchone()[0]
            print(f"[DEBUG] guardar_nuevo_cliente - Clientes actuales en habitación {habitacion_id}: {clientes_actuales}")

            total_personas = 1 + len(personas_adicionales)
            if clientes_actuales + total_personas > 20: # CHANGED FROM 4 TO 20
                print(f"[DEBUG] guardar_nuevo_cliente - Límite excedido: actuales={clientes_actuales}, intentando={total_personas}")
                if request.is_json:
                    return jsonify({"success": False, "error": f"La habitación excedería el límite máximo de 20 clientes (actuales: {clientes_actuales}, intentando agregar: {total_personas})"})
                flash('La habitación excedería el límite máximo de 20 clientes.', 'error')
                return redirect(url_for('agregar_cliente_habitacion', habitacion_id=habitacion_id))

            print("[DEBUG] guardar_nuevo_cliente - Insertando cliente principal")
            cur.execute("""INSERT INTO clientes (hora_ingreso, nombre, tipo_doc, numero_doc, telefono, procedencia, habitacion_id, check_in, check_out, valor, observacion) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (hora_ingreso_time, nombre, tipo_doc, numero_doc, telefono, procedencia, habitacion_id, check_in_dt, check_out_dt, valor, observacion))

            for persona in personas_adicionales:
                if persona.get('nombre'):
                    print(f"[DEBUG] guardar_nuevo_cliente - Insertando persona adicional: {persona['nombre']}")
                    cur.execute("""INSERT INTO clientes (hora_ingreso, nombre, tipo_doc, numero_doc, telefono, procedencia, habitacion_id, check_in, check_out, valor, observacion) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (hora_ingreso_time, persona['nombre'], persona.get('tipo_doc', 'C.c'), persona.get('numero_doc', ''), persona.get('telefono', ''), persona.get('procedencia', ''), habitacion_id, check_in_dt, check_out_dt, 0, ''))

            # FIX: Ensure the room status update happens AFTER the client insertion
            print(f"[DEBUG] guardar_nuevo_cliente - Actualizando estado de habitación {habitacion_id} a 'ocupada'")
            cur.execute("""UPDATE habitaciones SET estado = 'ocupada' WHERE id = %s AND usuario_id = %s""", (habitacion_id, user_id))
            conn.commit()
            print(f"[DEBUG] guardar_nuevo_cliente - Habitación {habitacion_id} actualizada a 'ocupada', commit realizado")

            if request.is_json:
                return jsonify({"success": True, "message": f"Cliente principal y {len(personas_adicionales)} persona(s) adicional(es) registrado(s) exitosamente"})
            else:
                flash('Cliente(s) registrado(s) exitosamente.', 'success')
                return redirect(url_for('index'))

    except pymysql.MySQLError as e:
        print(f"[DEBUG] guardar_nuevo_cliente - Error MySQL: {str(e)}")
        if request.is_json:
            return jsonify({"success": False, "error": f"Error en la base de datos: {str(e)}"})
        flash(f'Error en la base de datos: {str(e)}', 'error')
        return redirect(url_for('index'))
    except Exception as e:
        print(f"[DEBUG] guardar_nuevo_cliente - Error general: {str(e)}")
        if request.is_json:
            return jsonify({"success": False, "error": f"Error al guardar cliente: {str(e)}"})
        flash(f'Error al guardar nuevo cliente: {str(e)}', 'error')
        return redirect(url_for('index'))
    finally:
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()

# ----------------- REGISTRAR CLIENTE -----------------
@app.route('/registrar_cliente/<int:habitacion_id>', methods=['POST'])
def registrar_cliente(habitacion_id):
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))
    
    nombre = request.form.get('nombre')
    tipo_doc = request.form.get('tipo_doc')
    numero_doc = request.form.get('numero_doc')
    telefono = request.form.get('telefono')
    procedencia = request.form.get('procedencia')
    check_in = request.form.get('check_in')
    check_out_fecha = request.form.get('check_out')
    valor = request.form.get('valor')
    observacion = request.form.get('observacion')

    try:
        # Parse check-in datetime
        check_in_dt = datetime.strptime(check_in, "%Y-%m-%dT%H:%M")
        
        # Parse check-out date
        check_out_dt = datetime.strptime(check_out_fecha, "%Y-%m-%d")
        
        conn = get_db_connection()
        if not conn:
            flash('Error de conexión a la base de datos.', 'error')
            return redirect(url_for('index'))

        cur = conn.cursor()
        
        # Get user's configured checkout time
        cur.execute("SELECT checkout_hora FROM usuarios WHERE id = %s", (user_id,))
        checkout_config = cur.fetchone()
        
        if checkout_config and checkout_config[0]:
            checkout_time = checkout_config[0]
            
            # Handle different types returned by MySQL
            if isinstance(checkout_time, timedelta):
                # MySQL TIME type returns as timedelta
                total_seconds = int(checkout_time.total_seconds())
                checkout_hour = total_seconds // 3600
                checkout_minute = (total_seconds % 3600) // 60
            elif isinstance(checkout_time, str):
                # String format "HH:MM:SS" or "HH:MM"
                checkout_time_obj = datetime.strptime(checkout_time[:5], "%H:%M").time()
                checkout_hour = checkout_time_obj.hour
                checkout_minute = checkout_time_obj.minute
            elif isinstance(checkout_time, time):
                # Already a time object
                checkout_hour = checkout_time.hour
                checkout_minute = checkout_time.minute
            else:
                # Fallback to default
                checkout_hour = 13
                checkout_minute = 0
        else:
            # Default to 1 PM if not configured
            checkout_hour = 13
            checkout_minute = 0
        
        check_out_dt = datetime(check_out_dt.year, check_out_dt.month, check_out_dt.day, checkout_hour, checkout_minute)
        
        # Extract time for hora_ingreso field
        hora_ingreso_time = check_in_dt.time()
        
    except ValueError as e:
        flash(f'Error en el formato de fechas: {str(e)}', 'error')
        return redirect(url_for('index'))

    personas_adicionales = []
    form_keys = list(request.form.keys())
    
    # Extract additional persons data from form
    for key in form_keys:
        if key.startswith('personas_adicionales[') and key.endswith('][nombre]'):
            # Extract the index from the key
            import re
            match = re.search(r'personas_adicionales\[(\d+)\]\[nombre\]', key)
            if match:
                index = match.group(1)
                nombre_adicional = request.form.get(f'personas_adicionales[{index}][nombre]')
                if nombre_adicional:
                    personas_adicionales.append({
                        'nombre': nombre_adicional,
                        'tipo_doc': request.form.get(f'personas_adicionales[{index}][tipo_doc]', 'C.c'),
                        'numero_doc': request.form.get(f'personas_adicionales[{index}][numero_doc]', ''),
                        'telefono': request.form.get(f'personas_adicionales[{index}][telefono]', ''),
                        'procedencia': request.form.get(f'personas_adicionales[{index}][procedencia]', '')
                    })

    try:
        cur.execute("SELECT id FROM habitaciones WHERE id = %s AND usuario_id = %s", (habitacion_id, user_id))
        if not cur.fetchone():
            flash('No tienes permisos para modificar esta habitación.', 'error')
            return redirect(url_for('index'))
        
        cur.execute("""SELECT COUNT(*) FROM clientes WHERE habitacion_id = %s AND (check_out IS NULL OR check_out > NOW())""", (habitacion_id,))
        clientes_actuales = cur.fetchone()[0]

        total_personas = 1 + len(personas_adicionales)
        if clientes_actuales + total_personas > 4:
            flash(f'La habitación excedería el límite máximo de 4 clientes (actuales: {clientes_actuales}, intentando agregar: {total_personas}).', 'error')
            return redirect(url_for('agregar_cliente_habitacion', habitacion_id=habitacion_id))

        cur.execute("""INSERT INTO clientes (hora_ingreso, nombre, tipo_doc, numero_doc, telefono, procedencia, habitacion_id, check_in, check_out, valor, observacion) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (hora_ingreso_time, nombre, tipo_doc, numero_doc, telefono, procedencia, habitacion_id, check_in_dt, check_out_dt, valor, observacion))

        # Insert additional persons
        for persona in personas_adicionales:
            cur.execute("""INSERT INTO clientes (hora_ingreso, nombre, tipo_doc, numero_doc, telefono, procedencia, habitacion_id, check_in, check_out, valor, observacion) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (hora_ingreso_time, persona['nombre'], persona['tipo_doc'], persona['numero_doc'], persona['telefono'], persona['procedencia'], habitacion_id, check_in_dt, check_out_dt, 0, ''))

        cur.execute("""UPDATE habitaciones SET estado = 'ocupada' WHERE id = %s AND usuario_id = %s""", (habitacion_id, user_id))

        conn.commit()
        flash(f'Cliente principal y {len(personas_adicionales)} persona(s) adicional(es) registrado(s) exitosamente.', 'success')
        return redirect(url_for('index'))

    except pymysql.MySQLError as e:
        flash(f'Error en la base de datos: {str(e)}', 'error')
        return redirect(url_for('agregar_cliente_habitacion', habitacion_id=habitacion_id))
    finally:
        cur.close()
        conn.close()

# ----------------- LIBERAR -----------------
@app.route('/liberar/<int:habitacion_id>')
def liberar(habitacion_id):
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión a la base de datos.', 'error')
        return redirect(url_for('index'))

    try:
        cur = conn.cursor()
        
        cur.execute("SELECT numero FROM habitaciones WHERE id = %s AND usuario_id = %s", (habitacion_id, user_id))
        habitacion = cur.fetchone()
        if not habitacion:
            flash('No tienes permisos para liberar esta habitación.', 'error')
            return redirect(url_for('index'))
        
        # Intentar liberar clientes activos
        cur.execute("""UPDATE clientes SET check_out = NOW() WHERE habitacion_id = %s AND (check_out IS NULL OR check_out > NOW())""", (habitacion_id,))
        clientes_liberados = cur.rowcount

        # Siempre liberar la habitación, tenga o no clientes
        cur.execute("UPDATE habitaciones SET estado = 'libre' WHERE id = %s AND usuario_id = %s", (habitacion_id, user_id))

        if clientes_liberados > 0:
            flash(f'Habitación {habitacion[0]} y sus ocupantes liberados.')
        else:
            flash(f'Habitación {habitacion[0]} liberada (no había clientes activos).', 'warning')

        conn.commit()
        return redirect(url_for('index'))

    except pymysql.MySQLError as e:
        flash(f'Error en la base de datos: {str(e)}', 'error')
        return redirect(url_for('index'))
    finally:
        cur.close()
        conn.close()

# ----------------- EXPORTAR EXCEL -----------------
@app.route('/exportar_excel')
def exportar_excel():
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Error de conexión a la base de datos.', 'error')
        return redirect(url_for('index'))

    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT c.nombre, c.tipo_doc, c.numero_doc, c.telefono, c.procedencia, 
                   c.check_in, c.check_out, c.valor, c.observacion, h.numero AS habitacion_numero,
                   c.hora_ingreso
            FROM clientes c 
            JOIN habitaciones h ON c.habitacion_id = h.id 
            WHERE h.usuario_id = %s 
            ORDER BY c.check_in DESC
        """, (user_id,))
        all_clientes_data = cur.fetchall()

        if not all_clientes_data:
            flash('No hay datos para exportar', 'error')
            return redirect(url_for('index'))

        import os
        from openpyxl import Workbook
        from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime, date, time, timedelta

        desktop = os.path.join(os.path.expanduser("~"), "Desktop", "Nelson")
        os.makedirs(desktop, exist_ok=True)
        excel_path = os.path.join(desktop, "clientes_hotel.xlsx")

        # ✅ Encabezados actualizados
        columnas = [
            'Check-in', 'Habitación', 'Nombre', 'Tipo Doc', 'Número Doc', 'Teléfono', 
            'Procedencia', 'Check-out', 'Valor', 'Observación'
        ]

        wb = Workbook()
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color="FFFFFF")
        header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_style.alignment = Alignment(horizontal="center")

        if header_style.name not in wb.named_styles:
            wb.add_named_style(header_style)

        def format_time_value(tv):
            if tv is None:
                return None
            if isinstance(tv, time):
                return tv.strftime('%H:%M')
            if isinstance(tv, datetime):
                return tv.strftime('%H:%M')
            if isinstance(tv, timedelta):
                total_seconds = int(tv.total_seconds()) % (24 * 3600)
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                return f"{hours:02d}:{minutes:02d}"
            if isinstance(tv, str):
                if ':' in tv:
                    parts = tv.split(':')
                    try:
                        hh = int(parts[0]) % 24
                        mm = int(parts[1]) if len(parts) > 1 else 0
                        return f"{hh:02d}:{mm:02d}"
                    except Exception:
                        return tv
                return tv
            try:
                return str(tv)
            except Exception:
                return None

        def format_date_value(dv):
            if dv is None:
                return ''
            if isinstance(dv, datetime):
                return dv.strftime('%d/%m/%Y')
            if isinstance(dv, date):
                return dv.strftime('%d/%m/%Y')
            return str(dv)

        meses = {i: [] for i in range(1, 13)}
        for fila in all_clientes_data:
            check_in = fila[5]
            check_out = fila[6]
            hora_ingreso = fila[10] if len(fila) > 10 else None

            checkin_str = ''
            if isinstance(check_in, (datetime, date)):
                fecha_str = format_date_value(check_in)
                hora_str = format_time_value(hora_ingreso)
                if hora_str:
                    checkin_str = f"{fecha_str} {hora_str}"
                elif isinstance(check_in, datetime):
                    checkin_str = check_in.strftime('%d/%m/%Y %H:%M')
                else:
                    checkin_str = fecha_str
            else:
                checkin_str = str(check_in) if check_in else ''

            if isinstance(check_out, datetime):
                checkout_str = check_out.strftime('%d/%m/%Y %H:%M')
            elif isinstance(check_out, date):
                checkout_str = check_out.strftime('%d/%m/%Y')
            elif check_out is None:
                checkout_str = ''
            else:
                checkout_str = str(check_out)

            if isinstance(check_in, (datetime, date)):
                mes = check_in.month
            else:
                mes = datetime.now().month

            # ✅ Tipo y número de documento en celdas separadas
            nueva_fila = [
                checkin_str,                 # Check-in
                fila[9] or '',               # Habitación
                fila[0] or '',               # Nombre
                fila[1] or '',               # Tipo Doc
                fila[2] or '',               # Número Doc
                fila[3] or '',               # Teléfono
                fila[4] or '',               # Procedencia
                checkout_str,                # Check-out
                float(fila[7]) if fila[7] is not None else 0,  # Valor
                fila[8] or ''                # Observación
            ]
            meses[mes].append(nueva_fila)

        for mes, datos_mes in meses.items():
            if datos_mes:
                nombre_mes = [
                    'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
                ][mes - 1]

                ws = wb.create_sheet(title=nombre_mes)
                ws.append(columnas)

                for fila in datos_mes:
                    ws.append(fila)

                for cell in ws[1]:
                    cell.style = header_style

                for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
                    max_length = 0
                    col_letter = get_column_letter(col_idx)
                    for cell in column:
                        cell.border = thin_border
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = (max_length + 2) * 1.2
                    if adjusted_width > 0:
                        ws.column_dimensions[col_letter].width = adjusted_width

                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    ws.row_dimensions[row[0].row].height = 19.20

                ws.freeze_panes = 'A2'

        if 'Sheet' in wb.sheetnames and not wb['Sheet'].max_row > 1:
            del wb['Sheet']

        wb.save(excel_path)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_name = f'clientes_hotel_{timestamp}.xlsx'
        return send_file(excel_path, as_attachment=True, download_name=download_name)

    except pymysql.MySQLError as e:
        flash(f'Error en la base de datos: {str(e)}', 'error')
        return redirect(url_for('index'))
    except Exception as e:
        flash(f'Error al exportar Excel: {str(e)}', 'error')
        return redirect(url_for('index'))
    finally:
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()

# ----------------- CAMBIAR COLOR GENERAL -----------------
@app.route('/cambiar_color_general', methods=['POST'])
def cambiar_color_general():
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))

    habitacion_id_str = request.form.get('habitacion_id', '').strip()
    if not habitacion_id_str:
        flash("ID de habitación no válido", "error")
        return redirect(url_for('index'))

    try:
        habitacion_id = int(habitacion_id_str)
    except ValueError:
        flash("ID de habitación no válido", "error")
        return redirect(url_for('index'))

    nuevo_estado = request.form.get('nuevo_estado')
    nombre_cliente = request.form.get('nombre_cliente', '').strip()
    precio_noche = request.form.get('precio_noche', '0')
    
    if not nuevo_estado:
        flash('Datos incompletos para cambiar el estado.', 'error')
        return redirect(url_for('index'))
    
    if nuevo_estado == 'reservado':
        if not nombre_cliente:
            flash('El nombre del cliente es requerido para reservas.', 'error')
            return redirect(url_for('index'))
        
        try:
            precio_valor = float(precio_noche) if precio_noche else 0
        except ValueError:
            precio_valor = 0
            
        if precio_valor <= 0:
            flash('El precio por noche es requerido para reservas.', 'error')
            return redirect(url_for('index'))

    conn = get_db_connection()
    if not conn:
        flash("Error de conexión a la base de datos", "error")
        return redirect(url_for('index'))

    try:
        cur = conn.cursor()

        # Verificar que la habitación pertenezca al usuario
        cur.execute(
            "SELECT numero FROM habitaciones WHERE id = %s AND usuario_id = %s",
            (habitacion_id, user_id)
        )
        habit = cur.fetchone()
        if not habit:
            flash("No tienes permisos para modificar esta habitación", "error")
            return redirect(url_for('index'))

        if nuevo_estado == 'libre':
            # Liberar todos los clientes activos de la habitación
            cur.execute("""
                UPDATE clientes 
                SET check_out = NOW() 
                WHERE habitacion_id = %s 
                AND (check_out IS NULL OR check_out > NOW())
            """, (habitacion_id,))
            clientes_liberados = cur.rowcount
            
            # Actualizar el estado de la habitación a libre
            cur.execute(
                "UPDATE habitaciones SET estado = %s WHERE id = %s AND usuario_id = %s",
                (nuevo_estado, habitacion_id, user_id)
            )
            
            if clientes_liberados > 0:
                flash(f"Habitación {habit[0]} liberada y {clientes_liberados} cliente(s) dado(s) de baja", "success")
            else:
                flash(f"Habitación {habit[0]} liberada (no había clientes activos)", "success")
            
            conn.commit()
            return redirect(url_for('index'))

        # Actualizar el estado de la habitación
        cur.execute(
            "UPDATE habitaciones SET estado = %s WHERE id = %s AND usuario_id = %s",
            (nuevo_estado, habitacion_id, user_id)
        )

        if nuevo_estado == 'reservado' and nombre_cliente:
            try:
                precio_valor = float(precio_noche) if precio_noche else 0
            except ValueError:
                precio_valor = 0
                
            reserva_existente = verificar_reserva_existente(habitacion_id, user_id)
            
            if reserva_existente:
                datos_cliente = {
                    'nombre': nombre_cliente,
                    'tipo_doc': 'RESERVA',
                    'numero_doc': '',
                    'telefono': '',
                    'procedencia': '',
                    'check_in': datetime.now(),
                    'check_out': None,
                    'valor': precio_valor,
                    'observacion': f'Reserva actualizada desde panel - Precio: ${precio_valor}/noche'
                }
                
                if actualizar_reserva_existente(reserva_existente[0], datos_cliente):
                    flash(f"Reserva existente actualizada para habitación {habit[0]}", "success")
                else:
                    flash("Error al actualizar la reserva existente", "error")
            else:
                # Create new reservation if none exists
                cur.execute(
                    """INSERT INTO clientes (
                        hora_ingreso,
                        nombre,
                        tipo_doc,
                        numero_doc,
                        telefono,
                        procedencia,
                        habitacion_id,
                        check_in,
                        check_out,
                        valor,
                        observacion)
                       VALUES
                       (NOW(),
                        %s,
                        %s,    # tipo_doc por defecto
                        %s,    # numero_doc por defecto
                        %s,    # telefono por defecto
                        %s,    # procedencia por defecto
                        %s,
                        NOW(), # check_in con fecha/hora actual para evitar NULL
                        NULL,
                        %s,    # precio por noche
                        %s)""",
                    (
                        nombre_cliente,
                        'RESERVA',   # tipo_doc
                        '',          # numero_doc
                        '',          # telefono
                        '',          # procedencia
                        habitacion_id,
                        precio_valor, # valor (precio por noche)
                        f'Reserva creada desde panel - Precio: ${precio_valor}/noche'
                    )
                )
                flash(f"Nueva reserva creada para habitación {habit[0]}", "success")
        else:
            flash(f"Estado de la habitación {habit[0]} cambiado a {nuevo_estado.upper()}", "success")

        conn.commit()

    except pymysql.MySQLError as e:
        conn.rollback()
        flash(f"Error en la base de datos: {e}", "error")

    finally:
        cur.close()
        conn.close()

    return redirect(url_for('index'))

# ----------------- OBTENER DATOS HABITACION -----------------
@app.route('/obtener_datos_habitacion/<int:habitacion_id>')
def obtener_datos_habitacion(habitacion_id):
    user_id = require_user_session()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"})

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "error": "Error de conexión a la base de datos"})

    try:
        cur = conn.cursor()
        # Get current client data if exists
        cur.execute("""
            SELECT c.nombre, c.tipo_doc, c.numero_doc, c.telefono, c.procedencia, 
                   c.check_in, c.check_out, c.valor, c.observacion
            FROM clientes c 
            JOIN habitaciones h ON c.habitacion_id = h.id 
            WHERE h.id = %s AND h.usuario_id = %s 
            ORDER BY c.check_in DESC LIMIT 1
        """, (habitacion_id, user_id))
        
        cliente = cur.fetchone()
        
        if cliente:
            return jsonify({
                "success": True,
                "cliente": {
                    "nombre": cliente[0],
                    "tipo_doc": cliente[1],
                    "numero_doc": cliente[2],
                    "telefono": cliente[3],
                    "procedencia": cliente[4],
                    "check_in": cliente[5].strftime('%Y-%m-%dT%H:%M') if cliente[5] else '',
                    "check_out": cliente[6].strftime('%Y-%m-%d') if cliente[6] else '',
                    "valor": float(cliente[7]) if cliente[7] else 0,
                    "observacion": cliente[8] or ''
                }
            })
        else:
            return jsonify({"success": True, "cliente": None})
            
    except Exception as e:
        return jsonify({"success": False, "error": f"Error: {str(e)}"})
    finally:
        cur.close()
        conn.close()

@app.route('/reutilizar_ultimo', methods=['POST'])
def reutilizar_ultimo():
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    data = request.get_json()
    habitacion_id = data.get('habitacion_id')
    noches = data.get('noches', 1)
    precio_total = data.get('precio_total')
    nueva_fecha = data.get('nueva_fecha_ingreso')

    if not habitacion_id or not nueva_fecha:
        return jsonify({"success": False, "error": "Datos incompletos"}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "error": "Error de conexión a la base de datos"}), 500

    try:
        cur = conn.cursor()
        cur.execute("SET time_zone = '-05:00'")  # Forzar TZ MySQL

        cur.execute("""
            SELECT nombre, tipo_doc, numero_doc, telefono, procedencia, valor, observacion, check_in
            FROM clientes
            WHERE habitacion_id = %s
            ORDER BY check_in DESC
            LIMIT 10
        """, (habitacion_id,))
        huespedes = cur.fetchall()

        if not huespedes:
            return jsonify({"success": False, "error": "No hay huéspedes para reutilizar"}), 404

        # Agrupar por fecha de check_in para obtener todos los de la última estadía
        ultima_fecha_checkin = huespedes[0][7]  # check_in del primer resultado
        huespedes_ultima_estadia = [h for h in huespedes if h[7] == ultima_fecha_checkin]

        # Fecha nueva
        try:
            if 'T' in nueva_fecha:
                nueva_fecha_dt = datetime.strptime(nueva_fecha, "%Y-%m-%dT%H:%M")
            else:
                fecha_input = datetime.strptime(nueva_fecha, "%Y-%m-%d")
                now = datetime.now(ZoneInfo("America/Bogota"))
                nueva_fecha_dt = datetime.combine(fecha_input.date(), now.time())
        except ValueError:
            return jsonify({"success": False, "error": "Formato de fecha inválido"}), 400

        # Get user's configured checkout time
        cur.execute("SELECT checkout_hora FROM usuarios WHERE id = %s", (user_id,))
        checkout_config = cur.fetchone()
        
        if checkout_config and checkout_config[0]:
            checkout_time = checkout_config[0]
            
            # Handle different types returned by MySQL
            if isinstance(checkout_time, timedelta):
                # MySQL TIME type returns as timedelta
                total_seconds = int(checkout_time.total_seconds())
                checkout_hour = total_seconds // 3600
                checkout_minute = (total_seconds % 3600) // 60
            elif isinstance(checkout_time, str):
                # String format "HH:MM:SS" or "HH:MM"
                checkout_time_obj = datetime.strptime(checkout_time[:5], "%H:%M").time()
                checkout_hour = checkout_time_obj.hour
                checkout_minute = checkout_time_obj.minute
            elif isinstance(checkout_time, time):
                # Already a time object
                checkout_hour = checkout_time.hour
                checkout_minute = checkout_time.minute
            else:
                # Fallback to default
                checkout_hour = 13
                checkout_minute = 0
        else:
            # Default to 1 PM if not configured
            checkout_hour = 13
            checkout_minute = 0
        
        nueva_fecha_salida = nueva_fecha_dt.replace(hour=checkout_hour, minute=checkout_minute, second=0, microsecond=0) + timedelta(days=int(noches))
        hora_ingreso_time = nueva_fecha_dt

        for i, huesped in enumerate(huespedes_ultima_estadia):
            # El primer huésped lleva el valor total, los demás valor 0
            valor_final = precio_total if precio_total and i == 0 else (huesped[5] if i == 0 else 0)
            
            observaciones_existentes = huesped[6] if huesped[6] else ""
            nueva_observacion = f"{observaciones_existentes} | REUTILIZADO" if observaciones_existentes else "REUTILIZADO"

            cur.execute("""
            INSERT INTO clientes (hora_ingreso, habitacion_id, nombre, tipo_doc, numero_doc, telefono, 
                                  procedencia, check_in, check_out, valor, observacion)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (hora_ingreso_time, habitacion_id, huesped[0], huesped[1], huesped[2], 
                huesped[3], huesped[4], nueva_fecha_dt, nueva_fecha_salida, 
                valor_final, nueva_observacion))

        conn.commit()

        cur.execute("UPDATE habitaciones SET estado = 'ocupada' WHERE id = %s", (habitacion_id,))
        conn.commit()

        return jsonify({
            "success": True, 
            "message": f"Se reutilizaron {len(huespedes_ultima_estadia)} huésped(es) exitosamente"
        })

    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()

# ----------------- AGREGAR HABITACIÓN -----------------
@app.route('/agregar_habitacion', methods=['GET', 'POST'])
def agregar_habitacion():
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        numero = request.form['numero']
        descripcion = request.form['descripcion']
        estado = request.form['estado']

        conn = get_db_connection()
        if not conn:
            flash("Error de conexión a la base de datos", "error")
            return redirect(url_for('index'))

        try:
            cur = conn.cursor()
            
            cur.execute("SELECT id FROM habitaciones WHERE numero = %s AND usuario_id = %s", (numero, user_id))
            if cur.fetchone():
                flash(f"Ya tienes una habitación con el número {numero}. Cada habitación debe tener un número único en tu hotel.", "error")
                return render_template('agregar_habitacion.html')
            
            cur.execute("INSERT INTO habitaciones (numero, descripcion, estado, usuario_id) VALUES (%s, %s, %s, %s)", (numero, descripcion, estado, user_id))
            conn.commit()
            flash("Habitación agregada con éxito", "success")
        except Exception as e:
            flash(f"Error al agregar habitación: {e}", "error")
        finally:
            cur.close()
            conn.close()

        return redirect(url_for('index'))

    return render_template('agregar_habitacion.html')

# ----------------- ELIMINAR HABITACIÓN -----------------
@app.route('/eliminar_habitacion/<int:habitacion_id>', methods=['POST'])
def eliminar_habitacion(habitacion_id):
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    if not conn:
        flash("Error de conexión a la base de datos", "error")
        return redirect(url_for('index'))

    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM habitaciones WHERE id = %s AND usuario_id = %s", (habitacion_id, user_id))
        if cur.rowcount > 0:
            conn.commit()
            flash("Habitación eliminada con éxito")
        else:
            flash("No tienes permisos para eliminar esta habitación", "error")
    except pymysql.MySQLError as e:
        flash(f"Error al eliminar habitación: {str(e)}", "error")
    finally:
        cur.close()
        conn.close()
    
    return redirect(url_for('index'))

# ----------------- OBTENER HUESPEDES -----------------
@app.route('/obtener_huespedes/<int:habitacion_id>')
def obtener_huespedes(habitacion_id):
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "error": "Error de conexión a la base de datos"}), 500

    try:
        cur = conn.cursor(pymysql.cursors.DictCursor)
        # Verificar que la habitación pertenece a este usuario
        cur.execute(
            "SELECT id FROM habitaciones WHERE id = %s AND usuario_id = %s",
            (habitacion_id, user_id)
        )
        if not cur.fetchone():
            return jsonify({"success": False, "error": "No tienes permisos para acceder a esta habitación"})

        # Obtener huéspedes actuales
        cur.execute("""
            SELECT id, nombre, tipo_doc, numero_doc, telefono, procedencia,
                   check_in, check_out, valor, observacion
            FROM clientes
            WHERE habitacion_id = %s
              AND (check_out IS NULL OR check_out > NOW())
            ORDER BY (valor > 0) DESC, check_in DESC
        """, (habitacion_id,))
        filas = cur.fetchall()

        huespedes = [{
            'id': f['id'],
            'nombre': f['nombre'],
            'tipo_doc': f['tipo_doc'],
            'numero_doc': f['numero_doc'],
            'telefono': f['telefono'],
            'procedencia': f['procedencia'],
            'check_in': f['check_in'].isoformat() if f['check_in'] else None,
            'check_out': f['check_out'].isoformat() if f['check_out'] else None,
            'valor': float(f['valor']) if f['valor'] else 0,
            'observacion': f['observacion']
        } for f in filas]

        return jsonify({"success": True, "huespedes": huespedes})

    except Exception as e:
        return jsonify({"success": False, "error": f"Error al obtener huéspedes: {e}"}), 500
    finally:
        try:
            cur.close()
            conn.close()
        except:
            pass


# ----------------- RENUEVAR ESTADÍA -----------------
@app.route('/renovar_estadia', methods=['POST'])
def renovar_estadia():
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    
    try:
        data = request.get_json()
        habitacion_id = data.get('habitacion_id')
        tipo_renovacion = data.get('tipo_renovacion')
        dias_renovacion = data.get('dias_renovacion', 0)
        nueva_fecha_salida = data.get('nueva_fecha_salida')
        valor_adicional = data.get('valor_adicional', 0)
        observacion_renovacion = data.get('observacion_renovacion', '')
        
        if not habitacion_id:
            return jsonify({'success': False, 'error': 'ID de habitación requerido'}), 400
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Error de conexión a la base de datos'}), 500
        
        try:
            cur = conn.cursor()
            
            # Verify room belongs to user and get current client info
            cur.execute("""
                SELECT c.id, c.nombre, c.check_out, c.valor, c.observacion
                FROM clientes c 
                JOIN habitaciones h ON c.habitacion_id = h.id 
                WHERE h.id = %s AND h.usuario_id = %s 
                AND (c.check_out IS NULL OR c.check_out > NOW())
                AND c.valor > 0
                ORDER BY c.check_in DESC
                LIMIT 1
            """, (habitacion_id, user_id))
            
            cliente_info = cur.fetchone()
            if not cliente_info:
                return jsonify({'success': False, 'error': 'No se encontró cliente principal en esta habitación'}), 404
            
            cliente_id, nombre_cliente, fecha_salida_actual, valor_actual, observacion_actual = cliente_info
            
            # Calculate new checkout date
            if tipo_renovacion == 'dias':
                if fecha_salida_actual:
                    nueva_fecha_checkout = fecha_salida_actual + timedelta(days=dias_renovacion)
                else:
                    nueva_fecha_checkout = datetime.now() + timedelta(days=dias_renovacion)
                mensaje_renovacion = f"Cliente renovado por {dias_renovacion} día{'s' if dias_renovacion != 1 else ''}"
            else:  # tipo_renovacion == 'fecha'
                nueva_fecha_checkout = datetime.strptime(nueva_fecha_salida, '%Y-%m-%d')
                if fecha_salida_actual:
                    dias_diferencia = (nueva_fecha_checkout.date() - fecha_salida_actual.date()).days
                    mensaje_renovacion = f"Cliente renovado hasta {nueva_fecha_checkout.strftime('%d/%m/%Y')} ({dias_diferencia} días adicionales)"
                else:
                    mensaje_renovacion = f"Cliente renovado hasta {nueva_fecha_checkout.strftime('%d/%m/%Y')}"
            
            # Update client information
            nuevo_valor = valor_actual + valor_adicional if valor_adicional > 0 else valor_actual
            
            # Combine observations
            observaciones_combinadas = []
            if observacion_actual:
                observaciones_combinadas.append(observacion_actual)
            
            observaciones_combinadas.append(mensaje_renovacion)
            
            if observacion_renovacion:
                observaciones_combinadas.append(observacion_renovacion)
            
            nueva_observacion = " | ".join(observaciones_combinadas)
            
            # Update the client record
            cur.execute("""
                UPDATE clientes 
                SET check_out = %s, valor = %s, observacion = %s
                WHERE id = %s
            """, (nueva_fecha_checkout, nuevo_valor, nueva_observacion, cliente_id))
            
            conn.commit()
            
            return jsonify({
                'success': True, 
                'message': f'Estadía renovada exitosamente. {mensaje_renovacion}',
                'nueva_fecha_salida': nueva_fecha_checkout.strftime('%Y-%m-%d'),
                'nuevo_valor': nuevo_valor
            })
            
        except pymysql.MySQLError as e:
            conn.rollback()
            return jsonify({'success': False, 'error': f'Error en la base de datos: {str(e)}'}), 500
        finally:
            cur.close()
            conn.close()
            
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error interno: {str(e)}'}), 500
    

# ----------------- OBTENER ÚLTIMO CLIENTE DE UNA HABITACIÓN -----------------
@app.route('/ultimo_cliente/<int:habitacion_id>', methods=['GET'])
def ultimo_cliente(habitacion_id):
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "error": "Error de conexión a la base de datos"}), 500

    try:
        cur = conn.cursor(pymysql.cursors.DictCursor)
        cur.execute("SET time_zone = '-05:00'")

        # Verificar que la habitación existe
        cur.execute("""
            SELECT estado FROM habitaciones 
            WHERE id = %s AND usuario_id = %s
        """, (habitacion_id, user_id))
        
        habitacion = cur.fetchone()
        if not habitacion:
            return jsonify({"success": False, "error": "Habitación no encontrada"}), 404

        # Buscar el último grupo de clientes basado en created_at (más confiable que check_in)
        cur.execute("""
            SELECT id, check_in, check_out, valor, created_at
            FROM clientes
            WHERE habitacion_id = %s
            ORDER BY created_at DESC, id DESC
            LIMIT 1
        """, (habitacion_id,))
        ultimo_cliente = cur.fetchone()

        if not ultimo_cliente:
            return jsonify({"success": False, "error": "No hay clientes registrados"}), 404

        ultimo_created_at = ultimo_cliente['created_at']

        cur.execute("""
            SELECT id, nombre, tipo_doc, numero_doc, telefono, procedencia,
                   check_in, check_out, valor, observacion, hora_ingreso, created_at
            FROM clientes
            WHERE habitacion_id = %s 
            AND created_at >= DATE_SUB(%s, INTERVAL 1 MINUTE)
            AND created_at <= DATE_ADD(%s, INTERVAL 1 MINUTE)
            ORDER BY created_at DESC, id DESC
        """, (habitacion_id, ultimo_created_at, ultimo_created_at))

        filas = cur.fetchall()

        def to_iso_or_str(value):
            if isinstance(value, (datetime, date)):
                return value.isoformat()
            if value is not None:
                return str(value)
            return None

        clientes = []
        total_valor = 0
        for fila in filas:
            valor = float(fila['valor']) if fila.get('valor') is not None else 0
            total_valor += valor
            clientes.append({
                "id": fila.get('id'),
                "nombre": fila.get('nombre'),
                "tipo_doc": fila.get('tipo_doc'),
                "numero_doc": fila.get('numero_doc'),
                "telefono": fila.get('telefono'),
                "procedencia": fila.get('procedencia'),
                "check_in": to_iso_or_str(fila.get('check_in')),
                "check_out": to_iso_or_str(fila.get('check_out')),
                "valor": valor,
                "observacion": fila.get('observacion') or '',
                "hora_ingreso": to_iso_or_str(fila.get('hora_ingreso'))
            })
        
        # Calcular noches
        noches = 1
        if clientes and clientes[0]['check_in'] and clientes[0]['check_out']:
            try:
                check_in_dt = datetime.fromisoformat(clientes[0]['check_in'])
                check_out_dt = datetime.fromisoformat(clientes[0]['check_out'])
                if check_out_dt > check_in_dt:
                    noches = (check_out_dt - check_in_dt).days
                else:
                    noches = 1
            except Exception as e:
                print(f"Error calculando noches: {e}")
                noches = 1

        return jsonify({
            "success": True, 
            "clientes": clientes,
            "cliente_principal": clientes[0] if clientes else None,
            "total_huespedes": len(clientes),
            "noches": noches,
            "precio_total": total_valor
        })

    except Exception as e:
        print(f"[v0] Error en ultimo_cliente: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()

@app.route('/cancelar_reserva/<int:habitacion_id>', methods=['POST'])
def cancelar_reserva(habitacion_id):
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "error": "Error de conexión a la base de datos"}), 500

    try:
        cur = conn.cursor()
        
        # Verificar que la habitación pertenece al usuario
        cur.execute("""
            SELECT estado FROM habitaciones 
            WHERE id = %s AND usuario_id = %s
        """, (habitacion_id, user_id))
        
        habitacion = cur.fetchone()
        if not habitacion:
            return jsonify({"success": False, "error": "Habitación no encontrada"}), 404

        cur.execute("""
            DELETE FROM clientes 
            WHERE habitacion_id = %s 
            AND (check_in > NOW() OR check_out IS NULL OR check_out >= NOW())
        """, (habitacion_id,))
        
        registros_eliminados = cur.rowcount

        cur.execute("""
            UPDATE habitaciones 
            SET estado = 'libre' 
            WHERE id = %s AND usuario_id = %s
        """, (habitacion_id, user_id))

        conn.commit()
        
        return jsonify({
            "success": True, 
            "message": f"Reserva cancelada exitosamente. Se eliminaron {registros_eliminados} registro(s)"
        })

    except Exception as e:
        conn.rollback()
        print(f"[v0] Error en cancelar_reserva: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()

# ----------------- CHECK-IN -----------------
@app.route('/checkin', methods=['POST'])
def checkin():
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    data = request.get_json()
    habitacion_id = data.get('habitacion_id')
    noches = data.get('noches', 1)
    precio_total = data.get('precio_total')

    if not habitacion_id:
        return jsonify({"success": False, "error": "Datos incompletos"}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "error": "Error de conexión a la base de datos"}), 500

    try:
        cur = conn.cursor()
        cur.execute("SET time_zone = '-05:00'")  # Forzar TZ MySQL

        # Hora actual Bogotá
        now = datetime.now(ZoneInfo("America/Bogota"))

        check_in = now
        hora_ingreso = now.time()
        # Calcula la hora de salida a las 13:00 del día correspondiente
        
        # Get user's configured checkout time
        cur.execute("SELECT checkout_hora FROM usuarios WHERE id = %s", (user_id,))
        checkout_config = cur.fetchone()
        
        if checkout_config and checkout_config[0]:
            checkout_time = checkout_config[0]
            
            # Handle different types returned by MySQL
            if isinstance(checkout_time, timedelta):
                # MySQL TIME type returns as timedelta
                total_seconds = int(checkout_time.total_seconds())
                checkout_hour = total_seconds // 3600
                checkout_minute = (total_seconds % 3600) // 60
            elif isinstance(checkout_time, str):
                # String format "HH:MM:SS" or "HH:MM"
                checkout_time_obj = datetime.strptime(checkout_time[:5], "%H:%M").time()
                checkout_hour = checkout_time_obj.hour
                checkout_minute = checkout_time_obj.minute
            elif isinstance(checkout_time, time):
                # Already a time object
                checkout_hour = checkout_time.hour
                checkout_minute = checkout_time.minute
            else:
                # Fallback to default
                checkout_hour = 13
                checkout_minute = 0
        else:
            # Default to 1 PM if not configured
            checkout_hour = 13
            checkout_minute = 0
        
        check_out = check_in.replace(hour=checkout_hour, minute=checkout_minute, second=0, microsecond=0) + timedelta(days=int(noches))

        # Verificar si ya hay clientes activos en esta habitación
        cur.execute("""
            SELECT id FROM clientes 
            WHERE habitacion_id = %s 
            AND (check_out IS NULL OR check_out > NOW())
        """, (habitacion_id,))
        
        if cur.fetchone():
            return jsonify({"success": False, "error": "La habitación ya tiene clientes activos. No se puede realizar un nuevo check-in."}), 400

        # Insertar el nuevo cliente
        cur.execute("""
            INSERT INTO clientes (hora_ingreso, habitacion_id, nombre, tipo_doc, numero_doc, telefono,
                                  procedencia, check_in, check_out, valor, observacion, estado)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'activo')
        """, (hora_ingreso, habitacion_id, data.get('nombre'), data.get('tipo_doc'),
              data.get('numero_doc'), data.get('telefono'), data.get('procedencia'),
              check_in, check_out, precio_total, data.get('observacion')))

        conn.commit()

        # Actualizar el estado de la habitación a ocupada
        cur.execute("UPDATE habitaciones SET estado = 'ocupada' WHERE id = %s", (habitacion_id,))
        conn.commit()

        return jsonify({"success": True, "message": "Check-in registrado exitosamente"})

    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        conn.close()

def liberar_habitaciones_automaticamente():
    """
    Función que se ejecuta en un hilo separado para liberar habitaciones automáticamente
    cuando llega la hora de check-out configurada por cada usuario.
    """
    while True:
        try:
            time.sleep(60)  # Check every minute
            
            now = datetime.now()
            
            # Only proceed if we're at the start of a new minute
            if now.second < 5:  # Within first 5 seconds of the minute
                conn = get_db_connection()
                if conn:
                    cur = conn.cursor()
                    
                    # Get all users with configured checkout times
                    cur.execute("SELECT id, checkout_hora FROM usuarios WHERE checkout_hora IS NOT NULL")
                    usuarios = cur.fetchall()
                    
                    for usuario_id, checkout_hora in usuarios:
                        # Handle different types returned by MySQL
                        if isinstance(checkout_hora, timedelta):
                            # MySQL TIME type returns as timedelta
                            total_seconds = int(checkout_hora.total_seconds())
                            checkout_hour = total_seconds // 3600
                            checkout_minute = (total_seconds % 3600) // 60
                        elif isinstance(checkout_hora, str):
                            # String format "HH:MM:SS" or "HH:MM"
                            checkout_time_obj = datetime.strptime(str(checkout_hora)[:5], "%H:%M").time()
                            checkout_hour = checkout_time_obj.hour
                            checkout_minute = checkout_time_obj.minute
                        elif isinstance(checkout_hora, time):
                            # Already a time object
                            checkout_hour = checkout_hour.hour
                            checkout_minute = checkout_hour.minute
                        else:
                            continue  # Skip if invalid type
                        
                        # Check if current time matches the checkout time (within 1 minute window)
                        if now.hour == checkout_hour and now.minute == checkout_minute:
                            # Mark clients as checked out if their check_out time has passed for this user's rooms
                            cur.execute("""
                                UPDATE clientes c
                                INNER JOIN habitaciones h ON c.habitacion_id = h.id
                                SET c.check_out = NOW(), c.estado = 'finalizado'
                                WHERE h.usuario_id = %s
                                AND c.check_out <= NOW() 
                                AND (c.check_out IS NOT NULL)
                                AND c.estado = 'activo'
                            """, (usuario_id,))
                            
                            # Then, release rooms that have no active clients for this user
                            cur.execute("""
                                UPDATE habitaciones h 
                                SET h.estado = 'libre' 
                                WHERE h.usuario_id = %s
                                AND h.estado IN ('ocupada', 'mensualidad', 'reservado')
                                AND NOT EXISTS (
                                    SELECT 1 FROM clientes c 
                                    WHERE c.habitacion_id = h.id 
                                    AND (c.check_out IS NULL OR c.check_out > NOW())
                                    AND c.estado = 'activo'
                                )
                            """, (usuario_id,))
                            
                            habitaciones_liberadas = cur.rowcount
                            conn.commit()
                            print(f"[{now}] {habitaciones_liberadas} habitaciones liberadas automáticamente para usuario {usuario_id} a las {checkout_hour:02d}:{checkout_minute:02d}")
                    
                    cur.close()
                    conn.close()
                    
        except Exception as e:
            print(f"Error en liberación automática: {e}")

# ----------------- LIMPIEZA DE OBSERVACIONES -----------------
def limpiar_observaciones_semanales():
    """
    Elimina observaciones diarias si no han sido actualizadas en más de 7 días.
    Esta función se ejecutará una vez a la semana.
    """
    while True:
        try:
            # Esperar hasta el inicio de la semana (Lunes 00:00:00)
            ahora = datetime.now()
            proximo_lunes = ahora + timedelta(days=(7 - ahora.weekday()))
            proximo_lunes = proximo_lunes.replace(hour=0, minute=0, second=0, microsecond=0)
            tiempo_espera = (proximo_lunes - ahora).total_seconds()
            
            print(f"Esperando {tiempo_espera} segundos hasta el próximo lunes para limpiar observaciones...")
            time.sleep(tiempo_espera)

            conn = get_db_connection()
            if conn:
                cur = conn.cursor()
                # Eliminar observaciones que no han sido actualizadas en los últimos 7 días
                cur.execute("""
                    DELETE FROM observaciones_diarias 
                    WHERE fecha_actualizacion < NOW() - INTERVAL 7 DAY
                """)
                registros_eliminados = cur.rowcount
                conn.commit()
                cur.close()
                conn.close()
                print(f"[{datetime.now()}] Limpieza semanal de observaciones: {registros_eliminados} registros eliminados.")
            else:
                print("No se pudo conectar a la base de datos para la limpieza de observaciones.")
                
        except Exception as e:
            print(f"Error en la limpieza semanal de observaciones: {e}")
            # Esperar un tiempo antes de reintentar si hay un error
            time.sleep(3600) # Esperar 1 hora

# Start observations cleanup thread
observations_cleanup_thread = threading.Thread(target=limpiar_observaciones_semanales, daemon=True)
observations_cleanup_thread.start()
print("Hilo de limpieza de observaciones iniciado")

@app.route('/obtener_observaciones', methods=['GET'])
def obtener_observaciones():
    """Get all observations for the current user"""
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "error": "Error de conexión a la base de datos"}), 500

    try:
        cur = conn.cursor(pymysql.cursors.DictCursor)
        cur.execute("""
            SELECT dia_semana, observacion, fecha_actualizacion 
            FROM observaciones_diarias 
            WHERE usuario_id = %s
            ORDER BY FIELD(dia_semana, 'lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo')
        """, (user_id,))
        
        observaciones = cur.fetchall()
        
        # Convert to dict with day as key
        obs_dict = {}
        for obs in observaciones:
            obs_dict[obs['dia_semana']] = {
                'observacion': obs['observacion'],
                'fecha_actualizacion': obs['fecha_actualizacion'].isoformat() if obs['fecha_actualizacion'] else None
            }
        
        return jsonify({"success": True, "observaciones": obs_dict})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

@app.route('/guardar_observacion', methods=['POST'])
def guardar_observacion():
    """Save or update an observation for a specific day"""
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    try:
        data = request.get_json()
        dia_semana = data.get('dia_semana')
        observacion = data.get('observacion', '').strip()
        
        if not dia_semana:
            return jsonify({"success": False, "error": "Día de la semana requerido"}), 400
        
        # Validate day of week
        dias_validos = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']
        if dia_semana not in dias_validos:
            return jsonify({"success": False, "error": "Día de la semana inválido"}), 400
        
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "error": "Error de conexión a la base de datos"}), 500
        
        try:
            cur = conn.cursor()
            
            if observacion:
                # Insert or update observation
                cur.execute("""
                    INSERT INTO observaciones_diarias (usuario_id, dia_semana, observacion)
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE 
                        observacion = VALUES(observacion),
                        fecha_actualizacion = CURRENT_TIMESTAMP
                """, (user_id, dia_semana, observacion))
                
                conn.commit()
                return jsonify({"success": True, "message": f"Observación guardada para {dia_semana}"})
            else:
                # If observation is empty, delete it
                cur.execute("""
                    DELETE FROM observaciones_diarias 
                    WHERE usuario_id = %s AND dia_semana = %s
                """, (user_id, dia_semana))
                
                conn.commit()
                return jsonify({"success": True, "message": f"Observación eliminada para {dia_semana}"})
                
        except pymysql.MySQLError as e:
            conn.rollback()
            return jsonify({"success": False, "error": f"Error en la base de datos: {str(e)}"}), 500
        finally:
            cur.close()
            conn.close()
            
    except Exception as e:
        return jsonify({"success": False, "error": f"Error interno: {str(e)}"}), 500

@app.route('/eliminar_observacion', methods=['POST'])
def eliminar_observacion():
    """Delete an observation for a specific day"""
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    try:
        data = request.get_json()
        dia_semana = data.get('dia_semana')
        
        if not dia_semana:
            return jsonify({"success": False, "error": "Día de la semana requerido"}), 400
        
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "error": "Error de conexión a la base de datos"}), 500
        
        try:
            cur = conn.cursor()
            cur.execute("""
                DELETE FROM observaciones_diarias 
                WHERE usuario_id = %s AND dia_semana = %s
            """, (user_id, dia_semana))
            
            conn.commit()
            
            if cur.rowcount > 0:
                return jsonify({"success": True, "message": f"Observación eliminada para {dia_semana}"})
            else:
                return jsonify({"success": False, "error": "No se encontró observación para eliminar"}), 404
                
        except pymysql.MySQLError as e:
            conn.rollback()
            return jsonify({"success": False, "error": f"Error en la base de datos: {str(e)}"}), 500
        finally:
            conn.close()
            
    except Exception as e:
        return jsonify({"success": False, "error": f"Error interno: {str(e)}"}), 500

# ----------------- CALENDARIO -----------------
@app.route('/calendario')
def calendario():
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    if not conn:
        flash("Error de conexión a la base de datos")
        return render_template('calendario.html', habitaciones=[])

    try:
        cur = conn.cursor()
        # Obtener habitaciones del usuario
        cur.execute("SELECT id, numero, descripcion, estado FROM habitaciones WHERE usuario_id = %s ORDER BY numero", (user_id,))
        habitaciones = cur.fetchall()
        
        return render_template('calendario.html', habitaciones=habitaciones)
    except Exception as e:
        flash(f"Error: {str(e)}")
        return render_template('calendario.html', habitaciones=[])
    finally:
        if 'cur' in locals():
            cur.close()
        if conn:
            conn.close()

# 
@app.route('/guardar_reserva_calendario', methods=['POST'])
def guardar_reserva_calendario():
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    try:
        data = request.get_json()
        habitacion_id = data.get('habitacion_id')
        nombre_cliente = data.get('nombre_cliente')
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')
        precio_total = data.get('precio_total', 0)
        
        observacion = data.get('observacion', '') or f"Reserva desde calendario - {(datetime.strptime(fecha_fin, '%Y-%m-%d') - datetime.strptime(fecha_inicio, '%Y-%m-%d')).days} días"
        
        if not all([habitacion_id, nombre_cliente, fecha_inicio, fecha_fin]):
            return jsonify({"success": False, "error": "Todos los campos son requeridos"})

        # Convertir fechas
        fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")

        # Check-in 2 PM y Check-out 1 PM
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "error": "Error de conexión a la base de datos"})

        cur = conn.cursor()
        
        # Get user's configured checkout time
        cur.execute("SELECT checkout_hora FROM usuarios WHERE id = %s", (user_id,))
        checkout_config = cur.fetchone()
        
        if checkout_config and checkout_config[0]:
            checkout_time = checkout_config[0]
            
            # Handle different types returned by MySQL
            if isinstance(checkout_time, timedelta):
                # MySQL TIME type returns as timedelta
                total_seconds = int(checkout_time.total_seconds())
                checkout_hour = total_seconds // 3600
                checkout_minute = (total_seconds % 3600) // 60
            elif isinstance(checkout_time, str):
                # String format "HH:MM:SS" or "HH:MM"
                checkout_time_obj = datetime.strptime(checkout_time[:5], "%H:%M").time()
                checkout_hour = checkout_time_obj.hour
                checkout_minute = checkout_time_obj.minute
            elif isinstance(checkout_time, time):
                # Already a time object
                checkout_hour = checkout_time.hour
                checkout_minute = checkout_time.minute
            else:
                # Fallback to default
                checkout_hour = 13
                checkout_minute = 0
        else:
            # Default to 1 PM if not configured
            checkout_hour = 13
            checkout_minute = 0

        fecha_inicio_dt = fecha_inicio_dt.replace(hour=14, minute=0)
        fecha_fin_dt = fecha_fin_dt.replace(hour=checkout_hour, minute=checkout_minute)


        # Verificar que la habitación pertenezca al usuario
        cur.execute(
            "SELECT numero, estado FROM habitaciones WHERE id = %s AND usuario_id = %s",
            (habitacion_id, user_id)
        )
        habitacion = cur.fetchone()
        if not habitacion:
            return jsonify({"success": False, "error": "No tienes permisos para esta habitación"})

        fecha_hoy = datetime.now().date()
        if fecha_inicio_dt.date() <= fecha_hoy and habitacion[1] != 'libre':
            return jsonify({
                "success": False,
                "error": f"La habitación {habitacion[0]} no está disponible para reservas que empiecen hoy. Estado actual: {habitacion[1]}. Para reservas inmediatas, la habitación debe estar libre."
            })

        # Verificar conflictos en tabla reservas (no en clientes)
        cur.execute("""
            SELECT COUNT(*) FROM reservas r
            WHERE r.habitacion_id = %s
            AND r.estado != 'cancelada'
            AND (
                (r.fecha_inicio <= %s AND r.fecha_fin > %s) OR
                (r.fecha_inicio < %s AND r.fecha_fin >= %s) OR
                (r.fecha_inicio >= %s AND r.fecha_inicio < %s)
            )
        """, (
            habitacion_id,
            fecha_inicio_dt, fecha_inicio_dt,
            fecha_fin_dt, fecha_fin_dt,
            fecha_inicio_dt, fecha_fin_dt
        ))

        conflictos = cur.fetchone()[0]
        if conflictos > 0:
            return jsonify({
                "success": False,
                "error": f"Ya existe una reserva en esas fechas para la habitación {habitacion[0]}"
            })

        # Insertar en tabla reservas
        cur.execute("""
            INSERT INTO reservas (
                habitacion_id, usuario_id, nombre_cliente, fecha_inicio, fecha_fin, precio_total, observacion, estado
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            habitacion_id, user_id, nombre_cliente,
            fecha_inicio_dt, fecha_fin_dt, precio_total,
            observacion,
            "pendiente"
        ))

        if fecha_inicio_dt.date() <= fecha_hoy:
            # Si empieza hoy o ya empezó, actualizar a confirmada y marcar habitación ocupada
            cur.execute("UPDATE reservas SET estado = 'confirmada' WHERE habitacion_id = %s AND fecha_inicio = %s", (habitacion_id, fecha_inicio_dt))
            cur.execute("UPDATE habitaciones SET estado = 'ocupada' WHERE id = %s", (habitacion_id,))
            mensaje_estado = "ocupada (reserva activa)"
        else:
            # Si es futura, mantener libre hasta la fecha
            mensaje_estado = "libre (reserva futura programada)"

        conn.commit()

        return jsonify({
            "success": True,
            "message": f"Reserva creada exitosamente para {nombre_cliente} en habitación {habitacion[0]}. Estado: {mensaje_estado}"
        })

    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
        return jsonify({"success": False, "error": f"Error al guardar reserva: {str(e)}"})
    finally:
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()

@app.route('/eliminar_reserva_calendario', methods=['POST'])
def eliminar_reserva_calendario():
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "error": "Error de conexión a la base de datos"}), 500

    try:
        data = request.get_json()
        print(f"DEBUG: Datos JSON recibidos: {data}")
        reserva_id = data.get('reserva_id')
        
        if not reserva_id:
            return jsonify({"success": False, "error": "ID de reserva requerido"})
        
        cur = conn.cursor()
        
        # 1. BUSCAR Y VERIFICAR PERMISOS (reserva + usuario_id)
        cur.execute("""
            SELECT r.id, r.nombre_cliente, h.numero, r.fecha_inicio
            FROM reservas r
            JOIN habitaciones h ON r.habitacion_id = h.id
            WHERE r.id = %s AND h.usuario_id = %s
        """, (reserva_id, user_id))
        
        reserva_info = cur.fetchone()
        
        if not reserva_info:
            # Este es el error más probable si envías el ID, pero no te pertenece
            return jsonify({"success": False, "error": "No tienes permisos para eliminar esta reserva o el ID no existe"})
        
        reserva_id, nombre_cliente, numero_habitacion, fecha_inicio = reserva_info

        # 2. VERIFICAR QUE SEA UNA RESERVA FUTURA
        # Se compara solo la fecha, sin incluir la hora actual, para evitar problemas al inicio del día.
        # Si 'fecha_inicio' incluye hora, es mejor convertirla a solo fecha.
        
        hoy = datetime.now().date()
        fecha_reserva = fecha_inicio.date() if hasattr(fecha_inicio, 'date') else fecha_inicio

        if fecha_reserva <= hoy:
            # Este error ocurre si la fecha de inicio es HOY o en el pasado.
            return jsonify({
                "success": False, 
                "error": f"No se puede eliminar la reserva de {nombre_cliente}. La fecha de inicio ya pasó o es hoy."
            })

        # 3. ELIMINAR RESERVA
        cur.execute("DELETE FROM reservas WHERE id = %s", (reserva_id,))
        conn.commit()
        
        return jsonify({
            "success": True, 
            "message": f"Reserva futura de {nombre_cliente} en habitación {numero_habitacion} eliminada exitosamente"
        })

    except Exception as e:
        # Asegura el rollback si algo falla
        if conn:
            conn.rollback()
        # Reporta el error técnico en la consola del servidor
        print(f"ERROR al eliminar reserva: {str(e)}")
        # Devuelve un mensaje genérico al cliente por seguridad
        return jsonify({"success": False, "error": f"Error interno del servidor al eliminar: {str(e)}"})
        
    finally:
        # Cerrar siempre la conexión
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()

@app.route('/editar_reserva_calendario', methods=['POST'])
def editar_reserva_calendario():
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    try:
        data = request.get_json()
        reserva_id = data.get('cliente_id')  # El formulario envía cliente_id pero es realmente reserva_id
        nombre_cliente = data.get('nombre_cliente')
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')
        precio_total = data.get('precio_total', 0)
        estado = data.get('estado', 'pendiente')

        if not all([reserva_id, nombre_cliente, fecha_inicio, fecha_fin]):
            return jsonify({"success": False, "error": "Todos los campos son requeridos"})

        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "error": "Error de conexión a la base de datos"})

        cur = conn.cursor()

        # Verificar que la reserva pertenece al usuario y que sigue siendo futura
        cur.execute("""
            SELECT r.habitacion_id, r.fecha_inicio 
            FROM reservas r
            JOIN habitaciones h ON r.habitacion_id = h.id
            WHERE r.id = %s AND h.usuario_id = %s
        """, (reserva_id, user_id))
        
        reserva_info = cur.fetchone()
        if not reserva_info:
            return jsonify({"success": False, "error": "No tienes permisos para editar esta reserva"})
        
        habitacion_id, fecha_inicio_actual = reserva_info
        
        # Si la reserva ya empezó, no se puede editar desde calendario
        if fecha_inicio_actual <= datetime.now():
            return jsonify({
                "success": False, 
                "error": "No se puede editar esta reserva porque ya está en curso. Para modificarla, usa el panel principal."
            })

        # Convertir fechas
        fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")

        # Agregar hora de check-in (2 PM) y check-out (1 PM)
        # Get user's configured checkout time
        cur.execute("SELECT checkout_hora FROM usuarios WHERE id = %s", (user_id,))
        checkout_config = cur.fetchone()
        
        if checkout_config and checkout_config[0]:
            checkout_time = checkout_config[0]
            
            # Handle different types returned by MySQL
            if isinstance(checkout_time, timedelta):
                # MySQL TIME type returns as timedelta
                total_seconds = int(checkout_time.total_seconds())
                checkout_hour = total_seconds // 3600
                checkout_minute = (total_seconds % 3600) // 60
            elif isinstance(checkout_time, str):
                # String format "HH:MM:SS" or "HH:MM"
                checkout_time_obj = datetime.strptime(checkout_time[:5], "%H:%M").time()
                checkout_hour = checkout_time_obj.hour
                checkout_minute = checkout_time_obj.minute
            elif isinstance(checkout_time, time):
                # Already a time object
                checkout_hour = checkout_time.hour
                checkout_minute = checkout_time.minute
            else:
                # Fallback to default
                checkout_hour = 13
                checkout_minute = 0
        else:
            # Default to 1 PM if not configured
            checkout_hour = 13
            checkout_minute = 0

        fecha_inicio_dt = fecha_inicio_dt.replace(hour=14, minute=0)
        fecha_fin_dt = fecha_fin_dt.replace(hour=checkout_hour, minute=checkout_minute)

        # Actualizar la reserva
        cur.execute("""
            UPDATE reservas 
            SET nombre_cliente = %s, fecha_inicio = %s, fecha_fin = %s, precio_total = %s,
                estado = %s, observacion = CONCAT(COALESCE(observacion, ''), ' | EDITADO DESDE CALENDARIO')
            WHERE id = %s AND fecha_inicio > NOW()
        """, (nombre_cliente, fecha_inicio_dt, fecha_fin_dt, precio_total, estado, reserva_id))

        if cur.rowcount == 0:
            return jsonify({
                "success": False, 
                "error": "No se pudo editar. La reserva ya no es futura o no existe."
            })

        conn.commit()

        return jsonify({
            "success": True, 
            "message": f"Reserva de {nombre_cliente} actualizada exitosamente"
        })

    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
        return jsonify({"success": False, "error": f"Error al editar reserva: {str(e)}"})
    finally:
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()


@app.route('/obtener_reservas_calendario')
def obtener_reservas_calendario():
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "error": "Error de conexión"})

    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT r.id, r.nombre_cliente, r.fecha_inicio, r.fecha_fin, r.precio_total, 
                   r.estado, h.numero, h.id, r.observacion
            FROM reservas r
            JOIN habitaciones h ON r.habitacion_id = h.id
            WHERE h.usuario_id = %s
            AND r.estado != 'cancelada'
            AND r.fecha_fin >= NOW()
            ORDER BY r.fecha_inicio
        """, (user_id,))
        
        reservas = []
        for row in cur.fetchall():
            reservas.append({
                'reserva_id': row[0],
                'id': row[0],  # Added id field for consistency
                'nombre': row[1],
                'fecha_inicio': row[2].strftime('%Y-%m-%d %H:%M') if row[2] else '',
                'fecha_fin': row[3].strftime('%Y-%m-%d %H:%M') if row[3] else '',
                'valor': float(row[4]) if row[4] else 0,
                'estado': row[5],
                'habitacion': row[6],
                'habitacion_id': row[7],
                'observacion': row[8] if row[8] else ''
            })
        
        return jsonify({"success": True, "reservas": reservas})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    finally:
        if 'cur' in locals():
            cur.close()
        if conn:
            conn.close()

@app.route('/estadisticas_habitaciones')
def estadisticas_habitaciones():
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "error": "Error de conexión"})

    try:
        cur = conn.cursor()

        # Contar habitaciones por estado
        cur.execute("""
            SELECT estado, COUNT(*) as cantidad
            FROM habitaciones
            WHERE usuario_id = %s
            GROUP BY estado
        """, (user_id,))

        estadisticas = {}
        total_habitaciones = 0

        for row in cur.fetchall():
            estado, cantidad = row
            estadisticas[estado] = cantidad
            total_habitaciones += cantidad

        # Calcular habitaciones no disponibles
        no_disponibles = estadisticas.get('ocupada', 0) + estadisticas.get('reservado', 0) + estadisticas.get('mensualidad', 0) + estadisticas.get('mantenimiento', 0)

        return jsonify({
            "success": True,
            "estadisticas": {
                "libres": estadisticas.get('libre', 0),
                "ocupadas": estadisticas.get('ocupada', 0),
                "reservadas": estadisticas.get('reservado', 0),
                "mensualidad": estadisticas.get('mensualidad', 0),
                "mantenimiento": estadisticas.get('mantenimiento', 0),
                "total": total_habitaciones,
                "no_disponibles": no_disponibles
            }
        })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    finally:
        if 'cur' in locals():
            cur.close()
        if conn:
            conn.close()

# ----------------- CONFIGURACIÓN HORA CHECKOUT -----------------
@app.route('/obtener_config_checkout')
def obtener_config_checkout():
    """Obtiene la configuración de hora de checkout para el usuario"""
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "error": "Error de conexión a la base de datos"}), 500

    try:
        cur = conn.cursor()

        # Verificar si existe configuración para este usuario
        cur.execute("""
            SELECT checkout_hora
            FROM config_checkout
            WHERE usuario_id = %s
        """, (user_id,))

        config = cur.fetchone()

        if config:
            checkout_hora = config[0]
        else:
            # Si no existe configuración, usar valor por defecto
            checkout_hora = '13:00'

        return jsonify({
            "success": True,
            "checkout_hora": checkout_hora
        })

    except Exception as e:
        print(f"[DEBUG] Error obteniendo config checkout: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'cur' in locals():
            cur.close()
        if conn:
            conn.close()

@app.route('/guardar_config_checkout', methods=['POST'])
def guardar_config_checkout():
    """Guarda la configuración de hora de checkout para el usuario"""
    user_id = require_user_session_json()
    if not user_id:
        return jsonify({"success": False, "error": "Sesión expirada"}), 401

    try:
        data = request.get_json()
        checkout_hora = data.get('checkout_hora')

        if not checkout_hora:
            return jsonify({"success": False, "error": "Hora de checkout requerida"}), 400

        # Validar formato de hora (HH:MM)
        import re
        if not re.match(r'^([01]?[0-9]|2[0-3]):[0-5][0-9]$', checkout_hora):
            return jsonify({"success": False, "error": "Formato de hora inválido"}), 400

        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "error": "Error de conexión a la base de datos"}), 500

        try:
            cur = conn.cursor()

            # Insertar o actualizar la configuración
            cur.execute("""
                INSERT INTO config_checkout (usuario_id, checkout_hora)
                VALUES (%s, %s)
                ON DUPLICATE KEY UPDATE
                    checkout_hora = VALUES(checkout_hora)
            """, (user_id, checkout_hora))

            conn.commit()

            print(f"[DEBUG] Config checkout guardada: usuario {user_id}, hora {checkout_hora}")

            return jsonify({
                "success": True,
                "message": f"Configuración guardada exitosamente. Hora de checkout: {checkout_hora}"
            })

        except pymysql.MySQLError as e:
            conn.rollback()
            print(f"[DEBUG] Error MySQL guardando config checkout: {e}")
            return jsonify({"success": False, "error": f"Error en la base de datos: {str(e)}"}), 500
        finally:
            cur.close()
            conn.close()

    except Exception as e:
        print(f"[DEBUG] Error guardando config checkout: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# ----------------- EDITAR PERFIL -----------------
@app.route('/editar_perfil', methods=['GET', 'POST'])
def editar_perfil():
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))
    
    if request.method == 'GET':
        conn = get_db_connection()
        if not conn:
            flash("Error de conexión a la base de datos", "error")
            return redirect(url_for('index'))
        
        try:
            cur = conn.cursor()
            cur.execute("SELECT username FROM usuarios WHERE id = %s", (user_id,))
            user = cur.fetchone()
            
            if not user:
                flash("Usuario no encontrado", "error")
                return redirect(url_for('index'))
            
            return render_template('editar_perfil.html', username=user[0])
        except Exception as e:
            flash(f"Error al cargar perfil: {e}", "error")
            return redirect(url_for('index'))
        finally:
            cur.close()
            conn.close()
    
    # POST request - update profile
    username = request.form.get('username', '').strip()
    password_actual = request.form.get('password_actual', '')
    nueva_password = request.form.get('nueva_password', '')
    confirmar_password = request.form.get('confirmar_password', '')
    
    if not username:
        flash("El nombre de usuario es requerido", "error")
        return redirect(url_for('editar_perfil'))
    
    # If user wants to change password, validate current password and new password match
    if nueva_password or confirmar_password or password_actual:
        if not password_actual:
            flash("Debes ingresar tu contraseña actual para cambiarla", "error")
            return redirect(url_for('editar_perfil'))
        
        if not nueva_password:
            flash("Debes ingresar una nueva contraseña", "error")
            return redirect(url_for('editar_perfil'))
        
        if nueva_password != confirmar_password:
            flash("Las contraseñas nuevas no coinciden", "error")
            return redirect(url_for('editar_perfil'))
        
        if len(nueva_password) < 6:
            flash("La nueva contraseña debe tener al menos 6 caracteres", "error")
            return redirect(url_for('editar_perfil'))
    
    conn = get_db_connection()
    if not conn:
        flash("Error de conexión a la base de datos", "error")
        return redirect(url_for('editar_perfil'))
    
    try:
        cur = conn.cursor()
        
        # Check if username is already taken by another user
        cur.execute("SELECT id FROM usuarios WHERE username = %s AND id != %s", (username, user_id))
        if cur.fetchone():
            flash("El nombre de usuario ya está en uso", "error")
            return redirect(url_for('editar_perfil'))
        
        # If changing password, verify current password
        if nueva_password:
            cur.execute("SELECT password FROM usuarios WHERE id = %s", (user_id,))
            current_user = cur.fetchone()
            if not current_user or not check_password_hash(current_user[0], password_actual):
                flash("La contraseña actual es incorrecta", "error")
                return redirect(url_for('editar_perfil'))
        
        # Update profile
        if nueva_password:
            from werkzeug.security import generate_password_hash
            hashed_password = generate_password_hash(nueva_password)
            cur.execute("UPDATE usuarios SET username = %s, password = %s WHERE id = %s", 
                       (username, hashed_password, user_id))
            flash("Perfil y contraseña actualizados exitosamente", "success")
        else:
            cur.execute("UPDATE usuarios SET username = %s WHERE id = %s", (username, user_id))
            flash("Perfil actualizado exitosamente", "success")
        
        # Update session username
        session['usuario'] = username
        
        conn.commit()
        return redirect(url_for('index'))
        
    except Exception as e:
        flash(f"Error al actualizar perfil: {e}", "error")
        return redirect(url_for('editar_perfil'))
    finally:
        cur.close()
        conn.close()

def verificar_reserva_existente(habitacion_id, user_id):
    """Verifica si ya existe una reserva activa para la habitación"""
    conn = get_db_connection()
    if not conn:
        return None
    
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT id, nombre, telefono, observacion, check_in, check_out, valor, tipo_doc, numero_doc, procedencia 
            FROM clientes 
            WHERE habitacion_id = %s 
            AND (check_out IS NULL OR check_out > NOW()) 
            ORDER BY check_in DESC 
            LIMIT 1
        """, (habitacion_id,))
        
        reserva = cur.fetchone()
        return reserva
    except Exception as e:
        print(f"Error verificando reserva existente: {e}")
        return None
    finally:
        cur.close()
        conn.close()

def actualizar_reserva_existente(cliente_id, datos_cliente):
    """Actualiza una reserva existente con los datos completos del cliente"""
    conn = get_db_connection()
    if not conn:
        return False
    
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE clientes 
            SET nombre = %s, tipo_doc = %s, numero_doc = %s, telefono = %s, 
                procedencia = %s, check_in = %s, check_out = %s, valor = %s, observacion = %s
            WHERE id = %s
        """, (
            datos_cliente['nombre'],
            datos_cliente['tipo_doc'],
            datos_cliente['numero_doc'],
            datos_cliente['telefono'],
            datos_cliente['procedencia'],
            datos_cliente['check_in'],
            datos_cliente['check_out'],
            datos_cliente['valor'],
            datos_cliente['observacion'],
            cliente_id
        ))
        
        conn.commit()
        return True
    except Exception as e:
        print(f"Error actualizando reserva: {e}")
        conn.rollback()
        return False
    finally:
        cur.close()
        conn.close()


# ----------------- EDITAR HABITACIÓN -----------------
@app.route('/editar_habitacion/<int:habitacion_id>')
def editar_habitacion(habitacion_id):
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión a la base de datos.', 'error')
        return redirect(url_for('index'))

    try:
        cur = conn.cursor()
        cur.execute("SELECT id, numero, descripcion, estado FROM habitaciones WHERE id = %s AND usuario_id = %s", (habitacion_id, user_id))
        habitacion = cur.fetchone()

        if not habitacion:
            flash('Habitación no encontrada o no tienes permisos para acceder a ella.', 'error')
            return redirect(url_for('index'))

        return render_template('editar_habitacion.html', habitacion=habitacion)

    except pymysql.MySQLError as e:
        flash(f'Error en la base de datos: {str(e)}', 'error')
        return redirect(url_for('index'))
    finally:
        cur.close()
        conn.close()

@app.route('/actualizar_habitacion', methods=['POST'])
def actualizar_habitacion():
    user_id = require_user_session()
    if not user_id:
        return redirect(url_for('login'))
    
    habitacion_id = request.form.get('habitacion_id')
    numero = request.form.get('numero')
    descripcion = request.form.get('descripcion')
    estado = request.form.get('estado')

    conn = get_db_connection()
    if not conn:
        flash('Error de conexión a la base de datos.', 'error')
        return redirect(url_for('index'))

    try:
        cur = conn.cursor()
        
        # Verificar que la habitación pertenezca al usuario
        cur.execute("SELECT id FROM habitaciones WHERE id = %s AND usuario_id = %s", (habitacion_id, user_id))
        if not cur.fetchone():
            flash('No tienes permisos para editar esta habitación.', 'error')
            return redirect(url_for('index'))
        
        # Actualizar la habitación
        cur.execute("""
            UPDATE habitaciones 
            SET numero = %s, descripcion = %s, estado = %s 
            WHERE id = %s AND usuario_id = %s
        """, (numero, descripcion, estado, habitacion_id, user_id))
        
        conn.commit()
        flash('Habitación actualizada exitosamente.', 'success')
        return redirect(url_for('index'))

    except pymysql.MySQLError as e:
        flash(f'Error en la base de datos: {str(e)}', 'error')
        return redirect(url_for('index'))
    finally:
        cur.close()
        conn.close()


if __name__ == '__main__':
    print("Iniciando servidor Flask...")
    print("Probando conexion a base de datos...")

    try:
        import pymysql
        print("Modulo pymysql importado correctamente")
    except ImportError:
        print("pymysql no esta instalado")
        print("Instala con: pip install pymysql")

    test_conn = get_db_connection()
    if test_conn:
        print("Conexion a base de datos exitosa")
        test_conn.close()
    else:
        print("Error de conexion a base de datos")
        print("Visita http://localhost:5000/test-db para mas detalles")

    liberation_thread = threading.Thread(target=liberar_habitaciones_automaticamente, daemon=True)
    liberation_thread.start()

    # Start observations cleanup thread
    observations_cleanup_thread = threading.Thread(target=limpiar_observaciones_semanales, daemon=True)
    observations_cleanup_thread.start()
    print("Hilo de limpieza de observaciones iniciado")

    app.run(debug=True, host='0.0.0.0', port=5000)
